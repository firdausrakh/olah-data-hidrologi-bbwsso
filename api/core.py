from __future__ import annotations

import io
import os
import base64
import csv
import hashlib
import html
import json
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from flask import Flask
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIG
# ============================================================

try:
    import config
except ImportError:
    config = None


def get_config(name: str, default: Any = "") -> Any:
    """
    Prioritas:
    1. Environment Variable
    2. config.py
    3. default
    """

    value = os.environ.get(name)

    if value is not None:
        return value

    if config is not None:
        return getattr(config, name, default)

    return default


# ============================================================
# PROJECT PATH
# ============================================================

CURRENT_DIR = Path(__file__).resolve().parent
ROOT_DIR = CURRENT_DIR.parent
DATA_DIR = ROOT_DIR / "data"
WIB = timezone(timedelta(hours=7))


def now_wib_naive() -> datetime:
    """Current BBWS local time (WIB) as a naive datetime for vendor comparisons."""
    return datetime.now(WIB).replace(tzinfo=None)


def _vendor_metadata_paths(vendor: str, filename: str) -> tuple[Path, Path]:
    """Return runtime + repository paths for persistent vendor metadata.

    Local runs write back to data/<vendor>/ so the cache can be committed with
    the repository. Vercel writes only to /tmp and falls back to the repository
    seed on a cold instance.
    """
    source = DATA_DIR / vendor / filename
    runtime = (Path("/tmp") / f"bbws_{vendor}_{filename}") if os.environ.get("VERCEL") else source
    return runtime, source


def _load_vendor_metadata(vendor: str, filename: str, default: Any) -> Any:
    runtime, source = _vendor_metadata_paths(vendor, filename)
    for path in (runtime, source):
        if not path.exists():
            continue
        try:
            with path.open("r", encoding="utf-8") as fh:
                value = json.load(fh)
            return value
        except (OSError, json.JSONDecodeError):
            continue
    return default


def _save_vendor_metadata(vendor: str, filename: str, value: Any) -> None:
    runtime, _source = _vendor_metadata_paths(vendor, filename)
    try:
        runtime.parent.mkdir(parents=True, exist_ok=True)
        tmp = runtime.with_suffix(runtime.suffix + ".tmp")
        with tmp.open("w", encoding="utf-8") as fh:
            json.dump(value, fh, ensure_ascii=False, indent=2)
        tmp.replace(runtime)
    except Exception as exc:
        print(f"Gagal menyimpan metadata {vendor}/{filename}: {exc}")


# ============================================================
# CENTRALIZED STATION NAME ALIASES
# ============================================================

STATION_ALIASES_PATH = DATA_DIR / "station_aliases.json"


def _alias_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _load_station_aliases() -> dict[str, dict[str, str]]:
    try:
        with STATION_ALIASES_PATH.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        print(f"Gagal membaca data/station_aliases.json: {exc}")
        return {}

    if not isinstance(payload, dict):
        return {}

    out: dict[str, dict[str, str]] = {}
    for vendor in ("beacon", "higertech", "tatonas", "dashindo"):
        raw_vendor = payload.get(vendor, {})
        if not isinstance(raw_vendor, dict):
            continue
        out[vendor] = {
            _alias_text(station_id): _alias_text(alias)
            for station_id, alias in raw_vendor.items()
            if _alias_text(station_id) and _alias_text(alias)
        }
    return out


STATION_ALIASES = _load_station_aliases()


def station_alias(vendor: str, station_id: Any, fallback: Any = "") -> str:
    """Return the operator-facing station name from data/station_aliases.json."""
    vendor_key = _alias_text(vendor).casefold()
    station_key = _alias_text(station_id)
    fallback_name = _alias_text(fallback)
    return STATION_ALIASES.get(vendor_key, {}).get(station_key) or fallback_name or station_key


def _alias_station_seed(
    vendor: str,
    rows: Any,
    id_field: str,
    name_field: str = "name",
) -> Any:
    """Overlay aliases onto repository/runtime metadata without mutating the source list."""
    if not isinstance(rows, list):
        return rows
    result: list[Any] = []
    for raw in rows:
        if not isinstance(raw, dict):
            result.append(raw)
            continue
        row = dict(raw)
        station_id = row.get(id_field)
        row[name_field] = station_alias(vendor, station_id, row.get(name_field) or station_id)
        result.append(row)
    return result


def _calendar_month_boundary(dt: datetime, months: int) -> datetime:
    """First day of the month N calendar months after dt's current month."""
    absolute = dt.year * 12 + (dt.month - 1) + max(1, months)
    year, month0 = divmod(absolute, 12)
    return datetime(year, month0 + 1, 1)


def _split_datetime_month_chunks(start: datetime, end: datetime, months: int) -> list[tuple[datetime, datetime]]:
    """Split an interval into chunks spanning at most N calendar months."""
    if end < start:
        return []
    chunks: list[tuple[datetime, datetime]] = []
    cursor = start
    while cursor <= end:
        boundary = _calendar_month_boundary(cursor, months)
        chunk_end = min(end, boundary - timedelta(minutes=1))
        if chunk_end < cursor:
            chunk_end = min(end, cursor + timedelta(days=1) - timedelta(minutes=1))
        chunks.append((cursor, chunk_end))
        cursor = chunk_end + timedelta(minutes=1)
    return chunks


def _split_date_month_chunks(first_day: Any, last_day: Any, months: int) -> list[tuple[str, str]]:
    if last_day < first_day:
        return []
    out: list[tuple[str, str]] = []
    cursor = first_day
    while cursor <= last_day:
        boundary = _calendar_month_boundary(datetime(cursor.year, cursor.month, 1), months).date()
        chunk_end = min(last_day, boundary - timedelta(days=1))
        out.append((cursor.isoformat(), chunk_end.isoformat()))
        cursor = chunk_end + timedelta(days=1)
    return out


def _timeout_like(exc: Exception) -> bool:
    if isinstance(exc, (requests.Timeout, TimeoutError)):
        return True
    text = str(exc).lower()
    return any(token in text for token in ("timeout", "timed out", "read timed", "connect timed"))


# ============================================================
# APPLICATION CONFIG
# ============================================================

BASE_URL = str(
    get_config(
        "BBWS_BASE_URL",
        "https://bbwsso.monitoring4system.com",
    )
).rstrip("/")

LOGIN_URL = f"{BASE_URL}/login"
ANALISA_URL = f"{BASE_URL}/analisa"
SET_TOKEN_URL = f"{BASE_URL}/analisa/set_token"
DATAPOS_URL = f"{BASE_URL}/datapos"
DATA_CHUNK_URL = f"{BASE_URL}/analisa/data_chunk"

TIMEOUT = int(
    get_config("BBWS_TIMEOUT", 45)
)

PARAMETER_CACHE_TTL = int(
    get_config("PARAMETER_CACHE_TTL", 6 * 60 * 60)
)

# Beacon upstream has a hard 25-day ceiling. BEACON_CHUNK_DAYS is the new
# explicit name; MAX_QUERY_DAYS remains a backward-compatible fallback.
BEACON_CHUNK_DAYS = max(1, min(25, int(
    get_config("BEACON_CHUNK_DAYS", get_config("MAX_QUERY_DAYS", 25))
)))
MAX_QUERY_DAYS = BEACON_CHUNK_DAYS

# Higertech's upstream export is intrinsically calendar-month based. The value
# controls UI/progress grouping; each actual export remains one month.
HIGERTECH_CHUNK_MONTHS = max(1, min(12, int(
    get_config("HIGERTECH_CHUNK_MONTHS", 1)
)))

# Safe request concurrency. Each vendor keeps its own conservative limit so
# upstream telemetry servers are not flooded. Beacon parallelism is applied
# only to /analisa/data_chunk after one parameter-specific token is ready.
BEACON_PARALLEL_WORKERS = max(1, min(4, int(
    get_config("BEACON_PARALLEL_WORKERS", 3)
)))

# Pengolahan BBWS fast-token cache. A set_sensordash token is tied to one
# logger/parameter pair and can be reused for independent data_chunk periods
# while the authenticated Beacon session remains warm.
BEACON_PROCESS_TOKEN_TTL = max(30, min(15 * 60, int(
    get_config("BEACON_PROCESS_TOKEN_TTL", 5 * 60)
)))

HIGERTECH_PARALLEL_WORKERS = max(1, min(6, int(
    get_config("HIGERTECH_PARALLEL_WORKERS", 3)
)))

HIGERTECH_EXPORT_CACHE_TTL = int(
    get_config("HIGERTECH_EXPORT_CACHE_TTL", 15 * 60)
)

HIGERTECH_EXPORT_CACHE_MAX = max(1, int(
    get_config("HIGERTECH_EXPORT_CACHE_MAX", 12)
))

HIGERTECH_EXPORT_TIMEOUT = max(30, min(180, int(
    get_config("HIGERTECH_EXPORT_TIMEOUT", 120)
)))

# Pengolahan Higertech: prefer the vendor's lightweight native 5-minute JSON
# chart endpoint for daily/monthly/range requests, then let the existing
# frontend aggregation work from raw points. Long ranges fall back to XLSX.
HIGERTECH_CHART_DAY_WORKERS = max(1, min(12, int(
    get_config("HIGERTECH_CHART_DAY_WORKERS", 8)
)))
HIGERTECH_CHART_TIMEOUT = max(2.0, min(30.0, float(
    get_config("HIGERTECH_CHART_TIMEOUT", 8)
)))
HIGERTECH_CHART_MAX_DAYS = max(1, min(120, int(
    get_config("HIGERTECH_CHART_MAX_DAYS", 62)
)))
HIGERTECH_CHART_CACHE_TTL = max(60, min(24 * 60 * 60, int(
    get_config("HIGERTECH_CHART_CACHE_TTL", 6 * 60 * 60)
)))
HIGERTECH_CHART_TODAY_CACHE_TTL = max(15, min(10 * 60, int(
    get_config("HIGERTECH_CHART_TODAY_CACHE_TTL", 60)
)))
HIGERTECH_CHART_CACHE_MAX = max(16, min(1024, int(
    get_config("HIGERTECH_CHART_CACHE_MAX", 256)
)))

USERNAME = str(
    get_config("BEACON_USERNAME", "")
).strip()

PASSWORD = str(
    get_config("BEACON_PASSWORD", "")
)

USERNAME_FIELD = str(
    get_config("BEACON_USERNAME_FIELD", "username")
).strip() or "username"

PASSWORD_FIELD = str(
    get_config("BEACON_PASSWORD_FIELD", "password")
).strip() or "password"

def _parse_app_passwords(raw: Any) -> tuple[str, ...]:
    """Parse APP_PASSWORDS from JSON array, comma/semicolon/newline text, or a single value."""
    if raw is None:
        return ()
    if isinstance(raw, (list, tuple, set)):
        values = [str(item).strip() for item in raw]
    else:
        text = str(raw).strip()
        if not text:
            return ()
        values: list[str]
        if text.startswith("["):
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                parsed = None
            if isinstance(parsed, list):
                values = [str(item).strip() for item in parsed]
            else:
                values = [part.strip() for part in re.split(r"[;,\n]+", text)]
        else:
            values = [part.strip() for part in re.split(r"[;,\n]+", text)]

    # Preserve order, drop blanks and accidental duplicates.
    return tuple(dict.fromkeys(value for value in values if value))


APP_PASSWORDS = _parse_app_passwords(
    get_config("APP_PASSWORDS", get_config("APP_PASSWORD", ""))
)

SESSION_SECRET = str(
    get_config("SESSION_SECRET", "")
).strip()

# ============================================================
# HIGERTECH SOURCE CONFIG
# ============================================================

HIGERTECH_BASE_URL = str(
    get_config(
        "HIGERTECH_BASE_URL",
        "https://bbwsserayuopak.higertech.com",
    )
).rstrip("/")

HIGERTECH_USERNAME = str(
    get_config("HIGERTECH_USERNAME", "")
).strip()

HIGERTECH_PASSWORD = str(
    get_config("HIGERTECH_PASSWORD", "")
)

HIGERTECH_LOGIN_URL = f"{HIGERTECH_BASE_URL}/Account/Login"
HIGERTECH_DOWNLOAD_PAGE_URL = f"{HIGERTECH_BASE_URL}/DownloadData"
HIGERTECH_STATIONS_URL = f"{HIGERTECH_BASE_URL}/DownloadData/GetDatatableStation"
HIGERTECH_EXPORT_URL = f"{HIGERTECH_BASE_URL}/DownloadData/Export"
HIGERTECH_CHART_URL = f"{HIGERTECH_BASE_URL}/Station/GetChartDataAwlrArr"


# ============================================================
# DASHINDO SOURCE CONFIG
# ============================================================

DASHINDO_BASE_URL = str(
    get_config(
        "DASHINDO_BASE_URL",
        "http://202.180.30.82",
    )
).rstrip("/")

DASHINDO_SOCKET_URL = str(
    get_config(
        "DASHINDO_SOCKET_URL",
        "http://202.180.30.82:8000",
    )
).rstrip("/")

DASHINDO_USERNAME = str(
    get_config("DASHINDO_USERNAME", "")
).strip()

DASHINDO_PASSWORD = str(
    get_config("DASHINDO_PASSWORD", "")
)

# Keep a single Dashindo request below the Vercel function maxDuration.
# The Socket.IO server normally responds much faster; this is only a guard.
DASHINDO_WAIT_TIMEOUT = int(
    get_config("DASHINDO_WAIT_TIMEOUT", 45)
)

DASHINDO_PARALLEL_WORKERS = max(1, min(4, int(
    get_config("DASHINDO_PARALLEL_WORKERS", 3)
)))
DASHINDO_CHUNK_MONTHS = max(1, min(6, int(
    get_config("DASHINDO_CHUNK_MONTHS", 3)
)))
# Pengolahan uses native get_n_data (raw minute/sub-minute samples) instead of
# downloadcsv. CSV remains a reliability fallback.
DASHINDO_DIRECT_RAW_ENABLED = str(
    get_config("DASHINDO_DIRECT_RAW_ENABLED", "1")
).strip().lower() not in {"0", "false", "no", "off"}
DASHINDO_TZ_OFFSET_HOURS = 7


# ============================================================
# TATONAS SOURCE CONFIG
# ============================================================

# Fixed source settings: no Vercel environment variables are needed for these.
TATONAS_BASE_URL = "https://tatonas.co.id"

TATONAS_USERNAME = str(
    get_config("TATONAS_USERNAME", "")
).strip()

TATONAS_PASSWORD = str(
    get_config("TATONAS_PASSWORD", "")
)

TATONAS_PLANT = "028"

TATONAS_LOGIN_URL = f"{TATONAS_BASE_URL}/admin/p/login"
TATONAS_HOME_URL = f"{TATONAS_BASE_URL}/admin/p/home"
TATONAS_RAW_PAGE_URL = f"{TATONAS_BASE_URL}/admin/p/trs_local_trs_raw_detail_all"
TATONAS_HARDWARE_URL = f"{TATONAS_BASE_URL}/admin/p/trs_local_mst_hardware_lookup2"
TATONAS_SENSOR_CATALOG_URL = f"{TATONAS_BASE_URL}/admin/p/trs_local_mst_sensor_list2"
TATONAS_DATA_URL = f"{TATONAS_BASE_URL}/admin/p/trs_local_trs_raw_detail_list_all"

TATONAS_CHUNK_MONTHS = max(1, min(6, int(
    get_config("TATONAS_CHUNK_MONTHS", 3)
)))
TATONAS_PARALLEL_WORKERS = max(1, min(4, int(
    get_config("TATONAS_PARALLEL_WORKERS", 2)
)))


# ============================================================
# POSITIONS
# ============================================================

POSITIONS_PATH = ROOT_DIR / "data" / "beacon" / "positions.json"

if not POSITIONS_PATH.exists():
    POSITIONS_PATH = CURRENT_DIR / "data" / "beacon" / "positions.json"

if not POSITIONS_PATH.exists():
    raise FileNotFoundError(
        "data/beacon/positions.json tidak ditemukan. "
        f"Path yang diperiksa: {POSITIONS_PATH}"
    )

with POSITIONS_PATH.open(
    "r",
    encoding="utf-8",
) as f:
    FALLBACK_POSITIONS = json.load(f)


# ============================================================
# FLASK
# ============================================================

app = Flask(
    __name__,
    template_folder=str(ROOT_DIR / "templates"),
    static_folder=str(ROOT_DIR / "static"),
    static_url_path="/static",
)

app.config["JSON_AS_ASCII"] = False

if not SESSION_SECRET:
    SESSION_SECRET = os.environ.get(
        "VERCEL_URL",
        "local-bbws-pengolah-session-secret",
    )

app.secret_key = SESSION_SECRET

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = bool(
    os.environ.get("VERCEL")
)
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(
    hours=8
)


# ============================================================
# HIGERTECH ADAPTER
# ============================================================

HIGERTECH_LOCK = threading.RLock()
_HIGERTECH_SESSION: requests.Session | None = None
_HIGERTECH_SESSION_AT = 0.0
_HIGERTECH_STATION_SEED = _alias_station_seed(
    "higertech",
    _load_vendor_metadata("higertech", "positions.json", []),
    "deviceId",
)
_HIGERTECH_PARAMETER_CATALOG = _load_vendor_metadata("higertech", "parameter_catalog.json", {})
_HIGERTECH_STATION_CACHE: tuple[float, list[dict[str, str]]] | None = (
    (time.time(), _HIGERTECH_STATION_SEED)
    if isinstance(_HIGERTECH_STATION_SEED, list) and _HIGERTECH_STATION_SEED
    else None
)
_HIGERTECH_EXPORT_CACHE: dict[tuple[str, str, str], tuple[float, bytes]] = {}
_HIGERTECH_CHART_CACHE: dict[tuple[str, str], tuple[float, list[dict[str, Any]]]] = {}
_HIGERTECH_CHART_CACHE_LOCK = threading.RLock()
HIGERTECH_SESSION_TTL = 15 * 60
HIGERTECH_STATION_CACHE_TTL = 6 * 60 * 60


def _normalize_higertech_station_name(value: str) -> str:
    name = clean_text(value)
    if not name or name != name.upper():
        return name
    acronyms = {"AWLR", "ARR", "AWS", "TMA", "BBWS"}
    return " ".join(word if word in acronyms else word.capitalize() for word in name.split())


def _higertech_new_session() -> requests.Session:
    client = requests.Session()
    client.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/151.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    })
    return client


def _higertech_login(force: bool = False) -> requests.Session:
    global _HIGERTECH_SESSION, _HIGERTECH_SESSION_AT

    with HIGERTECH_LOCK:
        if (
            not force
            and _HIGERTECH_SESSION is not None
            and time.time() - _HIGERTECH_SESSION_AT < HIGERTECH_SESSION_TTL
        ):
            return _HIGERTECH_SESSION

        if (
            not HIGERTECH_USERNAME
            or not HIGERTECH_PASSWORD
            or HIGERTECH_USERNAME.startswith("ISI_")
            or HIGERTECH_PASSWORD.startswith("ISI_")
        ):
            raise RuntimeError(
                "HIGERTECH_USERNAME dan HIGERTECH_PASSWORD belum dikonfigurasi."
            )

        client = _higertech_new_session()
        response = client.get(HIGERTECH_LOGIN_URL, timeout=TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        form = soup.find("form")
        if form is None:
            raise RuntimeError("Form login Higertech tidak ditemukan.")

        payload: dict[str, str] = {}
        username_field = None
        password_field = None
        for tag in form.find_all("input"):
            name = tag.get("name")
            if not name:
                continue
            field_type = (tag.get("type") or "").lower()
            if field_type == "hidden":
                payload[name] = tag.get("value") or ""
            elif field_type == "password":
                password_field = name
            elif field_type not in {"submit", "button", "checkbox", "radio"}:
                low = name.lower()
                if any(k in low for k in ("username", "user", "email", "login")):
                    username_field = name

        username_field = username_field or "Username"
        password_field = password_field or "Password"
        payload[username_field] = HIGERTECH_USERNAME
        payload[password_field] = HIGERTECH_PASSWORD

        action = form.get("action") or "/Account/Login"
        login_action = urljoin(HIGERTECH_BASE_URL + "/", action)
        response = client.post(
            login_action,
            data=payload,
            headers={"Referer": HIGERTECH_LOGIN_URL},
            allow_redirects=True,
            timeout=TIMEOUT,
        )
        response.raise_for_status()

        check = client.get(
            HIGERTECH_DOWNLOAD_PAGE_URL,
            headers={"Referer": HIGERTECH_BASE_URL},
            allow_redirects=True,
            timeout=TIMEOUT,
        )
        check.raise_for_status()
        if "/account/login" in check.url.lower():
            raise RuntimeError("Login Higertech gagal. Periksa akun yang digunakan.")

        _HIGERTECH_SESSION = client
        _HIGERTECH_SESSION_AT = time.time()
        return client


def _higertech_datatable_payload(length: int = -1) -> dict[str, str]:
    data = {
        "draw": "1", "start": "0", "length": str(length), "type": "all",
        "search[value]": "", "search[regex]": "false",
        "order[0][column]": "3", "order[0][dir]": "asc",
    }
    cols = [
        ("", "", "false", "false"),
        ("name", "Name", "true", "true"),
        ("deviceId", "DeviceId", "true", "true"),
        ("type", "Type", "true", "true"),
        ("", "", "false", "false"),
    ]
    for i, (col_data, name, searchable, orderable) in enumerate(cols):
        data[f"columns[{i}][data]"] = col_data
        data[f"columns[{i}][name]"] = name
        data[f"columns[{i}][searchable]"] = searchable
        data[f"columns[{i}][orderable]"] = orderable
        data[f"columns[{i}][search][value]"] = ""
        data[f"columns[{i}][search][regex]"] = "false"
    return data


def higertech_stations(force: bool = False) -> list[dict[str, str]]:
    global _HIGERTECH_STATION_CACHE
    with HIGERTECH_LOCK:
        if (
            not force
            and _HIGERTECH_STATION_CACHE
            and time.time() - _HIGERTECH_STATION_CACHE[0] < HIGERTECH_STATION_CACHE_TTL
        ):
            return _HIGERTECH_STATION_CACHE[1]

        client = _higertech_login()
        headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": HIGERTECH_DOWNLOAD_PAGE_URL,
        }
        rows = []
        for length in (-1, 1000):
            response = client.post(
                HIGERTECH_STATIONS_URL,
                data=_higertech_datatable_payload(length),
                headers=headers,
                timeout=TIMEOUT,
            )
            if response.status_code in (401, 403) or "/account/login" in response.url.lower():
                client = _higertech_login(force=True)
                response = client.post(
                    HIGERTECH_STATIONS_URL,
                    data=_higertech_datatable_payload(length),
                    headers=headers,
                    timeout=TIMEOUT,
                )
            response.raise_for_status()
            result = response.json()
            rows = result.get("data") or []
            if rows or length == 1000:
                break

        result = []
        seen = set()
        for row in rows:
            upstream_name = _normalize_higertech_station_name(str(row.get("name") or ""))
            device_id = clean_text(str(row.get("deviceId") or ""))
            name = station_alias("higertech", device_id, upstream_name)
            station_type = clean_text(str(row.get("type") or ""))
            if not name or not device_id or not station_type:
                continue
            key = (name.upper(), device_id.upper(), station_type.upper())
            if key in seen:
                continue
            seen.add(key)
            result.append({
                "name": name,
                "deviceId": device_id,
                "type": station_type,
            })
        result.sort(key=lambda x: (x["name"].lower(), x["deviceId"].lower()))
        _HIGERTECH_STATION_CACHE = (time.time(), result)
        _save_vendor_metadata("higertech", "positions.json", result)
        catalog: dict[str, list[dict[str, Any]]] = {}
        for station in result:
            params: list[dict[str, Any]] = []
            stype = str(station.get("type") or "")
            if stype in {"ARR", "AWS", "AWLR_ARR"}:
                params.extend(higertech_parameters_for("rain"))
            if stype in {"AWLR", "AWLR_ARR"}:
                params.extend(higertech_parameters_for("tma"))
            catalog[str(station.get("deviceId") or "")] = params
        global _HIGERTECH_PARAMETER_CATALOG
        _HIGERTECH_PARAMETER_CATALOG = catalog
        _save_vendor_metadata("higertech", "parameter_catalog.json", catalog)
        return result


def _higertech_station_by_device(device_id: str) -> dict[str, str]:
    for station in higertech_stations():
        if station["deviceId"] == device_id:
            return station
    for station in higertech_stations(force=True):
        if station["deviceId"] == device_id:
            return station
    raise RuntimeError(f"Device Higertech {device_id} tidak ditemukan.")


def _month_keys_between(start_dt: datetime, end_dt: datetime) -> list[str]:
    y, m = start_dt.year, start_dt.month
    out = []
    while (y, m) <= (end_dt.year, end_dt.month):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m == 13:
            y += 1
            m = 1
    return out


def _parse_any_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    raw = clean_text(str(value))
    if not raw:
        return None
    candidates = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def _xlsx_rows(content: bytes) -> tuple[list[str], list[list[Any]]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best = None
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        if not values:
            continue
        for idx, row in enumerate(values[:25]):
            cells = [clean_text(str(v)) if v is not None else "" for v in row]
            joined = " ".join(cells).lower()
            score = sum(bool(re.search(p, joined)) for p in [
                r"tanggal|waktu|time|date", r"tinggi|muka air|water level|stage",
                r"curah|hujan|rain|precip", r"jam|hour",
            ])
            nonempty = sum(bool(c) for c in cells)
            candidate = (score, nonempty, -idx, ws, idx, values)
            if best is None or candidate[:3] > best[:3]:
                best = candidate
    if best is None:
        raise RuntimeError("File export Higertech tidak berisi tabel yang dapat dibaca.")
    _, _, _, ws, header_idx, values = best
    raw_header = list(values[header_idx])
    headers = []
    used = {}
    for i, value in enumerate(raw_header):
        name = clean_text(str(value)) if value is not None else f"Kolom {i+1}"
        if not name:
            name = f"Kolom {i+1}"
        used[name] = used.get(name, 0) + 1
        if used[name] > 1:
            name = f"{name} {used[name]}"
        headers.append(name)
    rows = []
    for row in values[header_idx + 1:]:
        vals = list(row[:len(headers)]) + [None] * max(0, len(headers) - len(row))
        if not any(v not in (None, "") for v in vals):
            continue
        rows.append(vals[:len(headers)])
    return headers, rows


def _find_time_col(headers: list[str], rows: list[list[Any]]) -> int:
    for i, h in enumerate(headers):
        if re.search(r"waktu|timestamp|datetime|tanggal|date|time", h, re.I):
            return i
    for i in range(len(headers)):
        hits = 0
        for row in rows[:20]:
            if i < len(row) and _parse_any_datetime(row[i]):
                hits += 1
        if hits >= 2:
            return i
    return 0


_HIGERTECH_MONTHS = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4,
    "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
    "september": 9, "oktober": 10, "november": 11, "desember": 12,
}


def _parse_higertech_datetime(value: Any) -> datetime | None:
    """Parse waktu export Higertech, mis. '01 Juli 2026 00.05'."""
    if isinstance(value, datetime):
        return value
    raw = clean_text(str(value)) if value is not None else ""
    if not raw:
        return None

    # Format resmi export per-5-menit Higertech memakai nama bulan Indonesia
    # dan pemisah titik untuk jam/menit.
    match = re.fullmatch(
        r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})(?:\s+(\d{1,2})[.:](\d{2})(?::(\d{2}))?)?",
        raw,
        re.I,
    )
    if match:
        day = int(match.group(1))
        month = _HIGERTECH_MONTHS.get(match.group(2).lower())
        year = int(match.group(3))
        hour = int(match.group(4) or 0)
        minute = int(match.group(5) or 0)
        second = int(match.group(6) or 0)
        if month:
            try:
                return datetime(year, month, day, hour, minute, second)
            except ValueError:
                return None

    return _parse_any_datetime(raw.replace(".", ":"))


def _find_header_index(headers: list[str], patterns: list[str]) -> int | None:
    for pattern in patterns:
        for i, header in enumerate(headers):
            if re.search(pattern, clean_text(header), re.I):
                return i
    return None


def _canonicalize_higertech_rows(
    headers: list[str],
    rows: list[list[Any]],
    parameter_id: str,
) -> tuple[list[str], list[list[Any]]]:
    """
    Ubah struktur XLSX Higertech menjadi dua kolom yang deterministik.

    Export AWLR_ARR per 5 menit yang diamati:
      Tanggal | Jam/Menit | TMA | Debit (m3/s) | Curah Hujan (mm)

    Frontend tidak lagi perlu menebak kolom waktu/nilai.
    """
    time_idx = _find_header_index(headers, [
        r"^jam\s*/?\s*menit$",
        r"^waktu$",
        r"timestamp",
        r"datetime",
    ])
    if time_idx is None:
        time_idx = _find_header_index(headers, [r"^tanggal$"])
    if time_idx is None:
        raise RuntimeError("Kolom waktu pada export Higertech tidak ditemukan.")

    if parameter_id == "tma":
        value_idx = _find_header_index(headers, [
            r"^tma$",
            r"tinggi\s*muka\s*air",
            r"water\s*level",
            r"^stage$",
        ])
        output_header = "Tinggi Muka Air"
    elif parameter_id == "rain":
        value_idx = _find_header_index(headers, [
            r"curah\s*hujan",
            r"rainfall",
            r"precip",
        ])
        output_header = "Curah Hujan"
    else:
        raise RuntimeError("Parameter Higertech tidak dikenali.")

    if value_idx is None:
        raise RuntimeError(
            f"Kolom nilai {output_header} pada export Higertech tidak ditemukan. "
            f"Header terbaca: {', '.join(headers)}"
        )

    result: list[list[Any]] = []
    for row in rows:
        if time_idx >= len(row) or value_idx >= len(row):
            continue
        dt = _parse_higertech_datetime(row[time_idx])
        if dt is None:
            continue
        value = row[value_idx]
        if value is None or clean_text(str(value)) == "":
            continue
        try:
            if isinstance(value, (int, float)):
                number = float(value)
            else:
                number = float(clean_text(str(value)).replace(",", "."))
        except (TypeError, ValueError):
            continue
        result.append([dt.strftime("%Y-%m-%d %H:%M:%S"), number])

    return ["Waktu", output_header], result


def _higertech_clone_authenticated_session() -> requests.Session:
    """Create a thread-local session using the cached authenticated cookies.

    requests.Session is not guaranteed to be thread-safe. Each parallel export
    therefore gets its own connection pool while reusing the login cookies from
    the warm Higertech session.
    """
    master = _higertech_login()
    client = _higertech_new_session()
    client.cookies.update(master.cookies)
    return client


def _higertech_export_month(
    station: dict[str, str],
    ym: str,
    *,
    isolated_session: bool = False,
) -> bytes:
    cache_key = (station["deviceId"], station["type"], ym)
    with HIGERTECH_LOCK:
        cached = _HIGERTECH_EXPORT_CACHE.get(cache_key)
        if cached and time.time() - cached[0] < HIGERTECH_EXPORT_CACHE_TTL:
            return cached[1]

    client = (
        _higertech_clone_authenticated_session()
        if isolated_session
        else _higertech_login()
    )
    payload = {
        "stationName": station["name"],
        "stationType": station["type"],
        "deviceId": station["deviceId"],
        "periode": "month",
        "selecedData": "minute",
        "filterPeriode": ym,
        "downloadFileType": "excel",
    }
    headers = {
        "Accept": "*/*",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": HIGERTECH_DOWNLOAD_PAGE_URL,
    }
    response = client.post(HIGERTECH_EXPORT_URL, data=payload, headers=headers, timeout=HIGERTECH_EXPORT_TIMEOUT)
    if response.status_code in (401, 403) or "/account/login" in response.url.lower():
        # Refresh the shared login once, then retry with a fresh isolated pool.
        _higertech_login(force=True)
        client = _higertech_clone_authenticated_session() if isolated_session else _higertech_login()
        response = client.post(HIGERTECH_EXPORT_URL, data=payload, headers=headers, timeout=HIGERTECH_EXPORT_TIMEOUT)
    response.raise_for_status()
    ctype = (response.headers.get("Content-Type") or "").lower()
    if "application/json" in ctype or "text/html" in ctype:
        raise RuntimeError("Higertech tidak mengembalikan file Excel untuk periode " + ym)

    content = response.content
    with HIGERTECH_LOCK:
        _HIGERTECH_EXPORT_CACHE[cache_key] = (time.time(), content)
        # Keep the opportunistic Vercel/local memory cache bounded.
        if len(_HIGERTECH_EXPORT_CACHE) > HIGERTECH_EXPORT_CACHE_MAX:
            oldest = min(_HIGERTECH_EXPORT_CACHE, key=lambda k: _HIGERTECH_EXPORT_CACHE[k][0])
            _HIGERTECH_EXPORT_CACHE.pop(oldest, None)
    return content


def _higertech_data_xlsx(
    device_id: str,
    dari: str,
    sampai: str,
    parameter_id: str,
    *,
    isolated_session: bool = False,
) -> tuple[list[str], list[list[Any]], dict[str, str]]:
    station = _higertech_station_by_device(device_id)
    start_dt = datetime.strptime(dari[:16], "%Y-%m-%d %H:%M")
    end_dt = datetime.strptime(sampai[:16], "%Y-%m-%d %H:%M")
    if end_dt < start_dt:
        raise RuntimeError("Tanggal akhir lebih kecil dari tanggal awal.")

    canonical_headers: list[str] = []
    merged_rows: list[list[Any]] = []

    month_keys = _month_keys_between(start_dt, end_dt)
    monthly_content: dict[str, bytes] = {}

    if len(month_keys) > 1 and HIGERTECH_PARALLEL_WORKERS > 1:
        workers = min(HIGERTECH_PARALLEL_WORKERS, len(month_keys))
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="higertech") as executor:
            futures = {
                executor.submit(
                    _higertech_export_month, station, ym, isolated_session=True
                ): ym
                for ym in month_keys
            }
            for future in as_completed(futures):
                ym = futures[future]
                try:
                    monthly_content[ym] = future.result()
                except RuntimeError as exc:
                    # A month outside the logger's actual recording lifetime can
                    # legitimately return no Excel file. Do not discard successful
                    # months just because the requested range starts/ends empty.
                    if "tidak mengembalikan file Excel" in str(exc):
                        continue
                    raise
    else:
        for ym in month_keys:
            try:
                monthly_content[ym] = _higertech_export_month(
                    station, ym, isolated_session=isolated_session
                )
            except RuntimeError as exc:
                if "tidak mengembalikan file Excel" in str(exc):
                    continue
                raise

    # Parse/merge in chronological order even when downloads completed out of order.
    for ym in month_keys:
        content = monthly_content.get(ym)
        if not content:
            continue
        raw_headers, raw_rows = _xlsx_rows(content)
        headers, rows = _canonicalize_higertech_rows(
            raw_headers, raw_rows, parameter_id
        )
        if not canonical_headers:
            canonical_headers = headers
        merged_rows.extend(rows)

    if not canonical_headers:
        return [], [], station

    filtered: list[list[Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in merged_rows:
        dt = _parse_any_datetime(row[0] if row else None)
        if dt is None or not (start_dt <= dt <= end_dt):
            continue
        key = (str(row[0]), str(row[1] if len(row) > 1 else ""))
        if key in seen:
            continue
        seen.add(key)
        filtered.append(row)

    filtered.sort(key=lambda row: str(row[0]))
    return canonical_headers, filtered, station



def _higertech_chart_cache_get(device_id: str, day_key: str) -> list[dict[str, Any]] | None:
    key = (str(device_id), str(day_key))
    now = time.time()
    today_key = now_wib_naive().date().isoformat()
    ttl = HIGERTECH_CHART_TODAY_CACHE_TTL if day_key == today_key else HIGERTECH_CHART_CACHE_TTL
    with _HIGERTECH_CHART_CACHE_LOCK:
        item = _HIGERTECH_CHART_CACHE.get(key)
        if not item:
            return None
        created_at, rows = item
        if now - created_at >= ttl:
            _HIGERTECH_CHART_CACHE.pop(key, None)
            return None
        return list(rows)


def _higertech_chart_cache_put(device_id: str, day_key: str, rows: list[dict[str, Any]]) -> None:
    key = (str(device_id), str(day_key))
    with _HIGERTECH_CHART_CACHE_LOCK:
        _HIGERTECH_CHART_CACHE[key] = (time.time(), list(rows))
        if len(_HIGERTECH_CHART_CACHE) > HIGERTECH_CHART_CACHE_MAX:
            oldest = min(_HIGERTECH_CHART_CACHE, key=lambda k: _HIGERTECH_CHART_CACHE[k][0])
            _HIGERTECH_CHART_CACHE.pop(oldest, None)


def _higertech_chart_day(device_id: str, day_key: str) -> list[dict[str, Any]]:
    cached = _higertech_chart_cache_get(device_id, day_key)
    if cached is not None:
        return cached

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": HIGERTECH_BASE_URL,
        "Referer": f"{HIGERTECH_BASE_URL}/Station",
    }
    payload = {
        "deviceId": str(device_id),
        "selectedTime": "minute",
        "filterDate": str(day_key),
    }

    def request_once(*, force_login: bool = False) -> requests.Response:
        if force_login:
            _higertech_login(force=True)
        client = _higertech_clone_authenticated_session()
        return client.post(
            HIGERTECH_CHART_URL,
            data=payload,
            headers=headers,
            timeout=HIGERTECH_CHART_TIMEOUT,
            allow_redirects=True,
        )

    response = request_once()
    if response.status_code in (401, 403) or "/account/login" in response.url.lower():
        response = request_once(force_login=True)
    response.raise_for_status()

    text = response.text.strip()
    if not text:
        rows: list[dict[str, Any]] = []
    else:
        ctype = (response.headers.get("Content-Type") or "").lower()
        if "text/html" in ctype or "/account/login" in response.url.lower():
            response = request_once(force_login=True)
            response.raise_for_status()
            text = response.text.strip()
            if not text:
                rows = []
            else:
                if "text/html" in (response.headers.get("Content-Type") or "").lower():
                    raise RuntimeError("Higertech mengembalikan HTML saat meminta chart 5 menit.")
                payload_json = response.json()
                rows = payload_json.get("data") or [] if isinstance(payload_json, dict) else []
        else:
            payload_json = response.json()
            rows = payload_json.get("data") or [] if isinstance(payload_json, dict) else []

    if not isinstance(rows, list):
        raise RuntimeError("Format data chart 5 menit Higertech tidak valid.")
    cleaned = [row for row in rows if isinstance(row, dict)]
    _higertech_chart_cache_put(device_id, day_key, cleaned)
    return cleaned


def _parse_higertech_chart_local_time(row: dict[str, Any]) -> datetime | None:
    raw = clean_text(str(row.get("readingAt") or ""))
    if raw:
        # Vendor emits local WIB wall-clock time with a misleading trailing Z.
        try:
            return datetime.fromisoformat(raw[:-1] if raw.endswith("Z") else raw).replace(tzinfo=None)
        except ValueError:
            pass
    raw_utc = clean_text(str(row.get("readingAtUtc") or ""))
    if raw_utc:
        try:
            utc_naive = datetime.fromisoformat(raw_utc[:-1] if raw_utc.endswith("Z") else raw_utc).replace(tzinfo=None)
            return utc_naive + timedelta(hours=7)
        except ValueError:
            pass
    return None


def _higertech_data_chart(
    device_id: str,
    dari: str,
    sampai: str,
    parameter_id: str,
) -> tuple[list[str], list[list[Any]], dict[str, str]]:
    station = _higertech_station_by_device(device_id)
    start_dt = datetime.strptime(dari[:16], "%Y-%m-%d %H:%M")
    end_dt = datetime.strptime(sampai[:16], "%Y-%m-%d %H:%M")
    if end_dt < start_dt:
        raise RuntimeError("Tanggal akhir lebih kecil dari tanggal awal.")

    span_days = (end_dt.date() - start_dt.date()).days + 1
    if span_days > HIGERTECH_CHART_MAX_DAYS:
        raise RuntimeError("Rentang terlalu panjang untuk fast chart Higertech.")

    if parameter_id == "tma":
        value_key = "waterLevel"
        output_header = "Tinggi Muka Air"
    elif parameter_id == "rain":
        value_key = "rainfall"
        output_header = "Curah Hujan"
    else:
        raise RuntimeError("Parameter Higertech tidak dikenali.")

    days: list[str] = []
    cursor = start_dt.date()
    while cursor <= end_dt.date():
        days.append(cursor.isoformat())
        cursor += timedelta(days=1)

    parts: dict[str, list[dict[str, Any]]] = {}
    workers = min(HIGERTECH_CHART_DAY_WORKERS, len(days) or 1)
    if workers > 1 and len(days) > 1:
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="higertech-chart") as executor:
            futures = {executor.submit(_higertech_chart_day, device_id, day): day for day in days}
            for future in as_completed(futures):
                day = futures[future]
                parts[day] = future.result()
    else:
        for day in days:
            parts[day] = _higertech_chart_day(device_id, day)

    merged: dict[str, float] = {}
    for day in days:
        for item in parts.get(day, []):
            dt = _parse_higertech_chart_local_time(item)
            if dt is None or dt < start_dt or dt > end_dt:
                continue
            raw_value = item.get(value_key)
            if raw_value is None or clean_text(str(raw_value)) == "":
                continue
            try:
                value = float(str(raw_value).strip().replace(",", "."))
            except (TypeError, ValueError):
                continue
            merged[dt.strftime("%Y-%m-%d %H:%M:%S")] = value

    rows = [[stamp, merged[stamp]] for stamp in sorted(merged)]
    return ["Waktu", output_header], rows, station


def higertech_data(
    device_id: str,
    dari: str,
    sampai: str,
    parameter_id: str,
    *,
    isolated_session: bool = False,
) -> tuple[list[str], list[list[Any]], dict[str, str]]:
    """Pengolahan Higertech fast path using native 5-minute JSON.

    The raw 5-minute values are returned unchanged to the existing processing
    pipeline, so hourly/daily aggregation remains application-owned. The proven
    XLSX export is kept as a fail-safe and for long ranges where hundreds of
    daily chart requests would be less efficient.
    """
    try:
        return _higertech_data_chart(device_id, dari, sampai, parameter_id)
    except Exception as chart_exc:
        # Preserve the established XLSX path for reliability. Long-range chart
        # rejection is expected and should silently choose the monthly export.
        try:
            return _higertech_data_xlsx(
                device_id,
                dari,
                sampai,
                parameter_id,
                isolated_session=isolated_session,
            )
        except Exception:
            # If both routes fail, the lightweight route usually carries the
            # most actionable auth/format error; for intentional long-range
            # fallback the XLSX exception is more useful.
            if "Rentang terlalu panjang" not in str(chart_exc):
                raise chart_exc
            raise

def higertech_parameters_for(data_type: str) -> list[dict[str, str]]:
    if data_type == "rain":
        return [{"id": "rain", "name": "Curah Hujan", "type": "rain"}]
    return [{"id": "tma", "name": "Tinggi Muka Air", "type": "tma"}]



# ============================================================
# DASHINDO ADAPTER
# ============================================================

class DashindoError(RuntimeError):
    pass


# Field comes from DATA_NAVIGATION on the Dashindo AWLR page.
# Mapping MUST use sensor ID because SOWL008 has two fields.
DASHINDO_FIELD_BY_ID: dict[str, str] = {
    "58": "tma",
    "60": "tma",
    "64": "tma1",
    "83": "tma1",
    "91": "tmaikn",
    "92": "tma1",
    "93": "tmaikr",
    "99": "tma1",
    "142": "tma1",
    "168": "tma",
    "171": "tma",
    "172": "tma2",
    "214": "tma1",
    "215": "tma1",
    "299": "tma",
    "300": "tma",
}

# Operator-facing names are centralized in data/station_aliases.json.

DASHINDO_LOCK = threading.RLock()
_DASHINDO_STATION_SEED = _alias_station_seed(
    "dashindo",
    _load_vendor_metadata("dashindo", "positions.json", []),
    "id",
)
_DASHINDO_PARAMETER_CATALOG = _load_vendor_metadata("dashindo", "parameter_catalog.json", {})
_DASHINDO_STATION_CACHE: tuple[float, list[dict[str, Any]]] | None = (
    (time.time(), _DASHINDO_STATION_SEED)
    if isinstance(_DASHINDO_STATION_SEED, list) and _DASHINDO_STATION_SEED
    else None
)
_DASHINDO_AUTH_COOKIES: dict[str, str] = {}
_DASHINDO_AUTH_AT = 0.0
DASHINDO_STATION_CACHE_TTL = 6 * 60 * 60
DASHINDO_SESSION_TTL = 15 * 60


def _dashindo_json_lenient(text: str) -> Any:
    """Parse JSON despite UTF-8 BOM or a PHP Notice/Warning prefix."""
    cleaned = text.replace("\ufeff", "").strip()
    try:
        return json.loads(cleaned)
    except Exception as first_error:
        decoder = json.JSONDecoder()
        for i, ch in enumerate(cleaned):
            if ch not in "[{":
                continue
            try:
                obj, _end = decoder.raw_decode(cleaned[i:])
                return obj
            except Exception:
                continue
        raise first_error


def _dashindo_new_session() -> requests.Session:
    client = requests.Session()
    client.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/139.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
    })
    return client


def _dashindo_credentials_ready() -> bool:
    return bool(
        DASHINDO_USERNAME
        and DASHINDO_PASSWORD
        and not DASHINDO_USERNAME.startswith("ISI_")
        and not DASHINDO_PASSWORD.startswith("ISI_")
    )


def _dashindo_login_fresh() -> requests.Session:
    if not _dashindo_credentials_ready():
        raise DashindoError(
            "DASHINDO_USERNAME dan DASHINDO_PASSWORD belum dikonfigurasi."
        )

    client = _dashindo_new_session()
    page = client.get(
        f"{DASHINDO_BASE_URL}/",
        headers={
            "Accept": (
                "text/html,application/xhtml+xml,application/xml;q=0.9,"
                "image/avif,image/webp,*/*;q=0.8"
            ),
            "Referer": f"{DASHINDO_BASE_URL}/",
            "Upgrade-Insecure-Requests": "1",
        },
        timeout=TIMEOUT,
    )
    page.raise_for_status()

    token_match = re.search(
        r"""const\s+token\s*=\s*['"]([^'"]+)['"]""",
        page.text,
        re.I,
    )
    if not token_match:
        raise DashindoError("Token login Dashindo tidak ditemukan.")

    password_hash = hashlib.sha256(
        DASHINDO_PASSWORD.encode("utf-8")
    ).hexdigest()

    response = client.post(
        f"{DASHINDO_BASE_URL}/API/login.php",
        data={
            "rememberme": "1",
            "username": DASHINDO_USERNAME,
            "password": password_hash,
            "token": token_match.group(1),
        },
        headers={
            "Accept": "*/*",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": DASHINDO_BASE_URL,
            "Referer": f"{DASHINDO_BASE_URL}/",
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=TIMEOUT,
        allow_redirects=False,
    )
    if response.status_code >= 400:
        raise DashindoError(
            f"Login Dashindo ditolak HTTP {response.status_code}."
        )

    # Follow the same browser warm-up so scadash_user_token is available.
    for path in ("/?r=1", "/loginroute", "/dashboard/awlr.php"):
        try:
            client.get(
                f"{DASHINDO_BASE_URL}{path}",
                headers={"Referer": f"{DASHINDO_BASE_URL}/"},
                timeout=TIMEOUT,
                allow_redirects=True,
            )
        except requests.RequestException:
            pass

    # Validate authenticated session. The PHP endpoint requires HTTP_REFERER.
    meta = client.get(
        f"{DASHINDO_BASE_URL}/dashboard/API/get-mqtt-awlr.php",
        headers={
            "Accept": "*/*",
            "Referer": f"{DASHINDO_BASE_URL}/dashboard/settings-awlr.php",
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=TIMEOUT,
    )
    meta.raise_for_status()

    payload = _dashindo_json_lenient(meta.text)
    header = payload.get("header", {}) if isinstance(payload, dict) else {}
    if int(header.get("code", 0) or 0) != 200:
        raise DashindoError(
            f"Sesi Dashindo belum valid: "
            f"{header.get('code')} - {header.get('detail')}"
        )
    return client


def _dashindo_login(force: bool = False) -> requests.Session:
    """Return a fresh HTTP pool carrying a cached authenticated cookie jar.

    This keeps Dashindo login reuse fast while avoiding concurrent use of a
    single requests.Session object by Flask requests / Engine.IO downloads.
    """
    global _DASHINDO_AUTH_COOKIES, _DASHINDO_AUTH_AT
    with DASHINDO_LOCK:
        if (
            not force
            and _DASHINDO_AUTH_COOKIES
            and time.time() - _DASHINDO_AUTH_AT < DASHINDO_SESSION_TTL
        ):
            client = _dashindo_new_session()
            client.cookies.update(_DASHINDO_AUTH_COOKIES)
            return client

        authenticated = _dashindo_login_fresh()
        _DASHINDO_AUTH_COOKIES = requests.utils.dict_from_cookiejar(authenticated.cookies)
        _DASHINDO_AUTH_AT = time.time()
        client = _dashindo_new_session()
        client.cookies.update(_DASHINDO_AUTH_COOKIES)
        try:
            authenticated.close()
        except Exception:
            pass
        return client


def _dashindo_metadata(
    client: requests.Session,
) -> dict[str, Any]:
    response = client.get(
        f"{DASHINDO_BASE_URL}/dashboard/API/get-mqtt-awlr.php",
        headers={
            "Accept": "*/*",
            "Referer": f"{DASHINDO_BASE_URL}/dashboard/settings-awlr.php",
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=TIMEOUT,
    )
    response.raise_for_status()
    payload = _dashindo_json_lenient(response.text)
    if not isinstance(payload, dict):
        raise DashindoError("Metadata Dashindo bukan object JSON.")

    header = payload.get("header", {})
    if int(header.get("code", 0) or 0) != 200:
        raise DashindoError(
            f"Metadata Dashindo gagal: "
            f"{header.get('code')} - {header.get('detail')}"
        )
    return payload


def dashindo_stations(
    force: bool = False,
) -> list[dict[str, Any]]:
    """Return 16 active AWLRs, using the v1.5.3 operator mapping."""
    global _DASHINDO_STATION_CACHE

    with DASHINDO_LOCK:
        if (
            not force
            and _DASHINDO_STATION_CACHE
            and time.time() - _DASHINDO_STATION_CACHE[0]
            < DASHINDO_STATION_CACHE_TTL
        ):
            return _DASHINDO_STATION_CACHE[1]

        client = _dashindo_login()
        payload = _dashindo_metadata(client)
        rows = payload.get("data")
        if not isinstance(rows, list):
            raise DashindoError(
                "Metadata Dashindo tidak memiliki array data."
            )

        items: list[dict[str, Any]] = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            try:
                if int(item.get("is_show", 0)) != 1:
                    continue
            except Exception:
                continue

            sensor_id = str(item.get("id", "")).strip()
            device = str(item.get("device", "")).strip()
            field = DASHINDO_FIELD_BY_ID.get(sensor_id, "")
            if not sensor_id or not device or not field:
                continue

            mapping = station_alias(
                "dashindo",
                sensor_id,
                str(item.get("keterangan") or device),
            )
            items.append({
                "id": sensor_id,
                "device": device,
                "field": field,
                "name": mapping,
                "source_name": str(
                    item.get("keterangan") or mapping
                ),
                "group": str(item.get("awlr_group") or ""),
                "unit": str(item.get("satuan") or "m"),
            })

        # Non-Irigasi A-Z first, then Irigasi A-Z.
        items.sort(
            key=lambda s: (
                1 if "irigasi" in s["name"].casefold() else 0,
                s["name"].casefold(),
                s["source_name"].casefold(),
            )
        )

        if not items:
            raise DashindoError(
                "Tidak ada AWLR Dashindo aktif yang memiliki mapping field."
            )

        _DASHINDO_STATION_CACHE = (time.time(), items)
        _save_vendor_metadata("dashindo", "positions.json", items)
        catalog = {
            str(station["id"]): [{
                "id": station["field"],
                "name": "Tinggi Muka Air",
                "unit": station.get("unit") or "m",
                "source_unit": station.get("unit") or "m",
                "type": "tma",
            }]
            for station in items
        }
        global _DASHINDO_PARAMETER_CATALOG
        _DASHINDO_PARAMETER_CATALOG = catalog
        _save_vendor_metadata("dashindo", "parameter_catalog.json", catalog)
        return items


def _dashindo_station(
    sensor_id: str,
) -> dict[str, Any]:
    station = next(
        (
            s
            for s in dashindo_stations()
            if str(s["id"]) == str(sensor_id)
        ),
        None,
    )
    if not station:
        raise DashindoError(
            f"Pos Dashindo ID {sensor_id} tidak ditemukan."
        )
    return station


class _DashindoEngineIO:
    """Engine.IO v4 + Socket.IO default namespace via HTTP long-polling."""

    RS = "\x1e"

    def __init__(
        self,
        http: requests.Session,
        trace: list[str],
    ) -> None:
        self.http = http
        self.trace = trace
        self.sid: str | None = None
        self.ping_interval_ms = 25000
        self.ping_timeout_ms = 20000

    def _log(self, text: str) -> None:
        # Vercel logs are useful for troubleshooting, but no secrets are logged.
        self.trace.append(text)
        print(f"[Dashindo] {text}", flush=True)

    def _params(
        self,
        include_sid: bool = True,
    ) -> dict[str, str]:
        params = {
            "EIO": "4",
            "transport": "polling",
            "t": str(int(time.time() * 1000)),
        }
        if include_sid and self.sid:
            params["sid"] = self.sid
        return params

    def _headers(
        self,
        post: bool = False,
    ) -> dict[str, str]:
        headers = {
            "Accept": "*/*",
            "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
            "Origin": DASHINDO_BASE_URL,
            "Referer": f"{DASHINDO_BASE_URL}/",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }
        if post:
            headers["Content-Type"] = "text/plain;charset=UTF-8"
        return headers

    def open(self) -> None:
        self._log("handshake EIO=4")
        response = self.http.get(
            f"{DASHINDO_SOCKET_URL}/socket.io/",
            params=self._params(include_sid=False),
            headers=self._headers(),
            timeout=min(TIMEOUT, DASHINDO_WAIT_TIMEOUT),
        )
        response.raise_for_status()
        text = response.text.lstrip("\ufeff").strip()

        if not text.startswith("0"):
            raise DashindoError(
                "Handshake Engine.IO Dashindo tidak valid. "
                f"Body awal: {text[:160]!r}"
            )

        try:
            payload = json.loads(text[1:])
        except Exception as exc:
            raise DashindoError(
                "Handshake Engine.IO Dashindo tidak dapat diparse."
            ) from exc

        sid = payload.get("sid")
        if not sid:
            raise DashindoError(
                "Handshake Engine.IO Dashindo tidak memiliki SID."
            )

        self.sid = str(sid)
        self.ping_interval_ms = int(
            payload.get("pingInterval", 25000)
        )
        self.ping_timeout_ms = int(
            payload.get("pingTimeout", 20000)
        )
        self._log("SID diterima")

    def post_raw(self, packet: str) -> None:
        if not self.sid:
            raise DashindoError(
                "Engine.IO Dashindo belum memiliki SID."
            )

        response = self.http.post(
            f"{DASHINDO_SOCKET_URL}/socket.io/",
            params=self._params(),
            data=packet.encode("utf-8"),
            headers=self._headers(post=True),
            timeout=min(TIMEOUT, DASHINDO_WAIT_TIMEOUT),
        )
        response.raise_for_status()

    def get_raw(
        self,
        timeout: int,
    ) -> str:
        if not self.sid:
            raise DashindoError(
                "Engine.IO Dashindo belum memiliki SID."
            )

        response = self.http.get(
            f"{DASHINDO_SOCKET_URL}/socket.io/",
            params=self._params(),
            headers=self._headers(),
            timeout=max(1, timeout),
        )
        response.raise_for_status()
        return response.text.lstrip("\ufeff")

    @classmethod
    def split_packets(
        cls,
        payload: str,
    ) -> list[str]:
        if not payload:
            return []
        return [
            packet
            for packet in payload.split(cls.RS)
            if packet
        ]

    @staticmethod
    def decode_event(
        packet: str,
    ) -> tuple[str, Any] | None:
        # Engine.IO message (4) + Socket.IO event (2).
        if not packet.startswith("42"):
            return None

        try:
            values = json.loads(packet[2:])
        except Exception:
            return None

        if (
            not isinstance(values, list)
            or not values
            or not isinstance(values[0], str)
        ):
            return None

        event = values[0]
        if len(values) == 1:
            data: Any = None
        elif len(values) == 2:
            data = values[1]
        else:
            data = values[1:]
        return event, data

    def send_event(
        self,
        event: str,
        *args: Any,
    ) -> None:
        packet = "42" + json.dumps(
            [event, *args],
            ensure_ascii=False,
            separators=(",", ":"),
        )
        self.post_raw(packet)

    def poll_until(
        self,
        wanted: set[str],
        deadline: float,
    ) -> tuple[str, Any]:
        while time.monotonic() < deadline:
            remaining = max(
                1,
                int(deadline - time.monotonic()),
            )
            raw = self.get_raw(
                timeout=min(remaining + 2, 30)
            )

            for packet in self.split_packets(raw):
                if packet == "2":
                    # Engine.IO ping -> pong.
                    self.post_raw("3")
                    continue
                if packet == "3":
                    continue
                if packet.startswith("1"):
                    raise DashindoError(
                        "Server Engine.IO Dashindo menutup koneksi."
                    )

                decoded = self.decode_event(packet)
                if decoded:
                    event, data = decoded
                    self._log(f"event {event}")
                    if event in wanted:
                        return event, data

        raise DashindoError(
            "Timeout menunggu event Dashindo: "
            + ", ".join(sorted(wanted))
        )

    def close(self) -> None:
        if not self.sid:
            return
        try:
            self.post_raw("41")
        except Exception:
            pass
        try:
            self.post_raw("1")
        except Exception:
            pass


def _dashindo_websocket_auth(
    client: requests.Session,
    key: str,
) -> Any:
    response = client.post(
        f"{DASHINDO_BASE_URL}/dashboard/API/websocket-auth.php",
        data={"s": key},
        headers={
            "Accept": "*/*",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": DASHINDO_BASE_URL,
            "Referer": f"{DASHINDO_BASE_URL}/dashboard/awlr.php",
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=TIMEOUT,
    )
    response.raise_for_status()

    payload = _dashindo_json_lenient(response.text)
    if (
        not isinstance(payload, dict)
        or "data" not in payload
    ):
        raise DashindoError(
            "websocket-auth.php Dashindo tidak mengembalikan data autentikasi."
        )
    return payload["data"]



def _dashindo_get_n_data(
    client: requests.Session,
    device: str,
    field: str,
    tss: str,
    tse: str,
) -> dict[str, Any]:
    """Fetch Dashindo raw minute/sub-minute telemetry directly as JSON.

    The vendor's own sensor page emits get_n_data(device, field, [tss, tse]) and
    receives n_data {times, values}. The supplied HAR confirms this contains the
    same raw samples as download_csv, but avoids CSV generation, Base64 transfer,
    decoding, and DictReader parsing.
    """
    trace: list[str] = []
    engine = _DashindoEngineIO(http=client, trace=trace)
    deadline = time.monotonic() + DASHINDO_WAIT_TIMEOUT

    try:
        engine.open()
        engine.post_raw("40")
        _event, ehlo = engine.poll_until({"ehlo"}, deadline)
        if not isinstance(ehlo, dict) or not ehlo.get("key"):
            raise DashindoError("Event ehlo Dashindo tidak memiliki key.")

        auth_data = _dashindo_websocket_auth(client, str(ehlo["key"]))
        engine.send_event("message", auth_data)
        engine.poll_until({"auth"}, deadline)

        engine.send_event("get_n_data", device, field, [tss, tse])
        _event, data = engine.poll_until({"n_data"}, deadline)
        if not isinstance(data, dict):
            raise DashindoError("Payload n_data Dashindo tidak valid.")
        times = data.get("times")
        values = data.get("values")
        if not isinstance(times, list) or not isinstance(values, list):
            raise DashindoError("Payload n_data Dashindo tidak memiliki times/values.")
        return data
    except requests.RequestException as exc:
        raise DashindoError(
            "Gagal berkomunikasi dengan Engine.IO Dashindo: "
            f"{exc}. Trace: {' -> '.join(trace)}"
        ) from exc
    finally:
        engine.close()


def _dashindo_n_data_rows(
    data: dict[str, Any],
    start_dt: datetime,
    end_dt: datetime,
) -> list[list[Any]]:
    """Normalize direct n_data while preserving Pengolahan's old time semantics.

    n_data and download_csv expose the same source timestamp strings. The legacy
    Pengolahan CSV adapter treated those strings as UTC-naive then converted +7
    hours to WIB. Keep exactly that transform here so this performance change
    does not silently shift existing processed results.
    """
    times = data.get("times")
    values = data.get("values")
    if not isinstance(times, list) or not isinstance(values, list):
        raise DashindoError("Payload n_data Dashindo tidak valid.")

    rows: list[list[Any]] = []
    seen: set[tuple[str, str]] = set()
    for raw_time, raw_value in zip(times, values):
        try:
            dt_utc = datetime.strptime(str(raw_time).strip(), "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            dt = dt_utc.astimezone(
                timezone(timedelta(hours=DASHINDO_TZ_OFFSET_HOURS))
            ).replace(tzinfo=None)
        except ValueError:
            continue
        if dt < start_dt or dt > end_dt:
            continue
        try:
            value = float(str(raw_value).strip().replace(",", "."))
        except (TypeError, ValueError):
            continue
        stamp = dt.strftime("%Y-%m-%d %H:%M:%S")
        key = (stamp, str(value))
        if key in seen:
            continue
        seen.add(key)
        rows.append([stamp, value])

    rows.sort(key=lambda row: row[0])
    return rows

def _dashindo_download_csv(
    client: requests.Session,
    device: str,
    field: str,
    tss: str,
    tse: str,
) -> bytes:
    trace: list[str] = []
    engine = _DashindoEngineIO(
        http=client,
        trace=trace,
    )
    deadline = time.monotonic() + DASHINDO_WAIT_TIMEOUT

    try:
        engine.open()

        # Browser sends body "40" to connect the default Socket.IO namespace.
        engine.post_raw("40")

        _event, ehlo = engine.poll_until(
            {"ehlo"},
            deadline,
        )
        if (
            not isinstance(ehlo, dict)
            or not ehlo.get("key")
        ):
            raise DashindoError(
                "Event ehlo Dashindo tidak memiliki key."
            )

        auth_data = _dashindo_websocket_auth(
            client,
            str(ehlo["key"]),
        )
        engine.send_event(
            "message",
            auth_data,
        )
        engine.poll_until(
            {"auth"},
            deadline,
        )

        engine.send_event(
            "downloadcsv",
            device,
            field,
            tss,
            tse,
        )
        _event, data = engine.poll_until(
            {"download_csv"},
            deadline,
        )

        if not isinstance(data, dict):
            raise DashindoError(
                "Payload download_csv Dashindo tidak valid."
            )
        if not data.get("content"):
            # Empty periods are normal for ranges before installation / after
            # the latest telemetry. Return a valid empty CSV instead of turning
            # the whole user request into Failed to fetch.
            return b"id,_field,_time,_value\n"

        try:
            return base64.b64decode(
                data["content"]
            )
        except Exception as exc:
            raise DashindoError(
                "Content CSV Dashindo bukan Base64 valid."
            ) from exc

    except requests.RequestException as exc:
        raise DashindoError(
            "Gagal berkomunikasi dengan Engine.IO Dashindo: "
            f"{exc}. Trace: {' -> '.join(trace)}"
        ) from exc
    finally:
        engine.close()


def _dashindo_csv_rows(
    raw_csv: bytes,
    expected_device: str,
    expected_field: str,
    start_dt: datetime,
    end_dt: datetime,
) -> list[list[Any]]:
    """Parse raw Dashindo CSV and keep repo processing unchanged."""
    text = raw_csv.decode(
        "utf-8-sig",
        errors="replace",
    )
    reader = csv.DictReader(
        io.StringIO(text)
    )

    required = {
        "id",
        "_field",
        "_time",
        "_value",
    }
    if (
        not reader.fieldnames
        or not required.issubset(
            set(reader.fieldnames)
        )
    ):
        raise DashindoError(
            "Format CSV Dashindo tidak sesuai "
            "(id,_field,_time,_value)."
        )

    rows: list[list[Any]] = []
    for item in reader:
        if (
            str(item.get("id", "")).strip()
            != expected_device
        ):
            continue
        if (
            str(item.get("_field", "")).strip()
            != expected_field
        ):
            continue

        raw_time = str(
            item.get("_time", "")
        ).strip()
        try:
            # Dashindo CSV timestamps are UTC-naive. Convert them explicitly to
            # WIB before filtering/returning so all downstream processing sees
            # local BBWS Serayu Opak time.
            dt_utc = datetime.strptime(
                raw_time,
                "%Y-%m-%d %H:%M:%S",
            ).replace(tzinfo=timezone.utc)
            dt = dt_utc.astimezone(
                timezone(timedelta(hours=DASHINDO_TZ_OFFSET_HOURS))
            ).replace(tzinfo=None)
        except ValueError:
            continue

        if dt < start_dt or dt > end_dt:
            continue

        raw_value = str(
            item.get("_value", "")
        ).strip()
        try:
            value: Any = float(raw_value)
        except ValueError:
            continue

        # Preserve Dashindo's familiar raw column names so the existing
        # repo auto-mapping / hourly processing handles it exactly like
        # an uploaded Dashindo CSV.
        rows.append([
            dt.strftime("%Y-%m-%d %H:%M:%S"),
            value,
        ])

    rows.sort(
        key=lambda row: row[0]
    )
    return rows


def dashindo_parameters_for(
    sensor_id: str,
    data_type: str,
) -> list[dict[str, Any]]:
    if data_type != "tma":
        return []

    cached = _DASHINDO_PARAMETER_CATALOG.get(str(sensor_id)) if isinstance(_DASHINDO_PARAMETER_CATALOG, dict) else None
    if isinstance(cached, list) and cached:
        return cached

    station = _dashindo_station(sensor_id)
    params = [{
        "id": station["field"],
        "name": "Tinggi Muka Air",
        "unit": station.get("unit") or "m",
        "source_unit": station.get("unit") or "m",
        "type": "tma",
    }]
    _DASHINDO_PARAMETER_CATALOG[str(sensor_id)] = params
    _save_vendor_metadata("dashindo", "parameter_catalog.json", _DASHINDO_PARAMETER_CATALOG)
    return params


def dashindo_data(
    sensor_id: str,
    field: str,
    dari: str,
    sampai: str,
) -> tuple[list[str], list[list[Any]], dict[str, Any], dict[str, Any]]:
    station = _dashindo_station(sensor_id)
    if field != station["field"]:
        raise DashindoError("Parameter Dashindo tidak sesuai dengan pos yang dipilih.")

    try:
        start_dt = datetime.strptime(dari[:16], "%Y-%m-%d %H:%M")
        end_dt = datetime.strptime(sampai[:16], "%Y-%m-%d %H:%M")
    except ValueError as exc:
        raise DashindoError("Format periode Dashindo tidak valid.") from exc
    if end_dt < start_dt:
        raise DashindoError("Tanggal akhir tidak boleh sebelum tanggal awal.")

    start_utc = start_dt - timedelta(hours=DASHINDO_TZ_OFFSET_HOURS)
    end_utc = end_dt - timedelta(hours=DASHINDO_TZ_OFFSET_HOURS)
    chunks = _split_date_month_chunks(start_utc.date(), end_utc.date(), DASHINDO_CHUNK_MONTHS)

    def fetch_period(period: tuple[str, str], depth: int = 0) -> list[list[Any]]:
        last_exc: Exception | None = None
        for attempt in range(2):
            try:
                client = _dashindo_login()
                if DASHINDO_DIRECT_RAW_ENABLED:
                    try:
                        data = _dashindo_get_n_data(
                            client=client, device=station["device"], field=station["field"],
                            tss=period[0], tse=period[1],
                        )
                        return _dashindo_n_data_rows(
                            data=data, start_dt=start_dt, end_dt=end_dt,
                        )
                    except Exception:
                        # Reliability fallback: preserve the proven raw CSV path.
                        # Both paths carry raw minute/sub-minute data; aggregation
                        # remains in the existing Pengolahan pipeline.
                        pass

                raw = _dashindo_download_csv(
                    client=client, device=station["device"], field=station["field"],
                    tss=period[0], tse=period[1],
                )
                return _dashindo_csv_rows(
                    raw_csv=raw, expected_device=station["device"], expected_field=station["field"],
                    start_dt=start_dt, end_dt=end_dt,
                )
            except Exception as exc:
                last_exc = exc
                if attempt == 0:
                    time.sleep(0.15)
        assert last_exc is not None
        a = datetime.strptime(period[0], "%Y-%m-%d").date()
        b = datetime.strptime(period[1], "%Y-%m-%d").date()
        days = (b - a).days
        if _timeout_like(last_exc) and depth < 3 and days >= 2:
            mid = a + timedelta(days=days // 2)
            left = fetch_period((a.isoformat(), mid.isoformat()), depth + 1)
            right_start = mid + timedelta(days=1)
            right = fetch_period((right_start.isoformat(), b.isoformat()), depth + 1)
            return left + right
        raise last_exc

    parts: dict[int, list[list[Any]]] = {}
    errors: dict[int, Exception] = {}
    if len(chunks) > 1 and DASHINDO_PARALLEL_WORKERS > 1:
        with ThreadPoolExecutor(
            max_workers=min(DASHINDO_PARALLEL_WORKERS, len(chunks)),
            thread_name_prefix="dashindo",
        ) as executor:
            futures = {executor.submit(fetch_period, period): idx for idx, period in enumerate(chunks)}
            for future in as_completed(futures):
                idx = futures[future]
                try:
                    parts[idx] = future.result()
                except Exception as exc:
                    errors[idx] = exc
    else:
        for idx, period in enumerate(chunks):
            try:
                parts[idx] = fetch_period(period)
            except Exception as exc:
                errors[idx] = exc

    if errors:
        edge_indexes = {0, len(chunks) - 1}
        non_edge = [idx for idx in errors if idx not in edge_indexes]
        if non_edge or not parts:
            detail = "; ".join(f"{chunks[i][0]}..{chunks[i][1]}: {err}" for i, err in sorted(errors.items()))
            raise DashindoError(f"Sebagian rentang Dashindo gagal: {detail}")

    rows: list[list[Any]] = []
    seen_rows: set[tuple[str, str]] = set()
    for idx in sorted(parts):
        for row in parts[idx]:
            key = (str(row[0]), str(row[1] if len(row) > 1 else ""))
            if key in seen_rows:
                continue
            seen_rows.add(key)
            rows.append(row)
    rows.sort(key=lambda row: str(row[0]))

    parameter = {
        "id": station["field"], "name": "Tinggi Muka Air",
        "unit": station.get("unit") or "m", "source_unit": station.get("unit") or "m",
        "type": "tma",
    }
    return ["_time", "_value"], rows, station, parameter


# ============================================================
# TATONAS ADAPTER
# ============================================================

TATONAS_LOCK = threading.RLock()
_TATONAS_SESSION: requests.Session | None = None
_TATONAS_SESSION_AT = 0.0
_TATONAS_CSRF = ""
_TATONAS_STATION_SEED = _alias_station_seed(
    "tatonas",
    _load_vendor_metadata("tatonas", "positions.json", []),
    "kd_hardware",
)
_TATONAS_PARAMETER_SEED = _load_vendor_metadata("tatonas", "parameter_catalog.json", {})
_TATONAS_STATION_CACHE: tuple[float, list[dict[str, Any]]] | None = (
    (time.time(), _TATONAS_STATION_SEED)
    if isinstance(_TATONAS_STATION_SEED, list) and _TATONAS_STATION_SEED
    else None
)
_TATONAS_SENSOR_CATALOG_CACHE: tuple[float, list[dict[str, Any]]] | None = None
_TATONAS_PARAMETER_CACHE: dict[str, tuple[float, list[dict[str, Any]]]] = {
    str(hw): (time.time(), params)
    for hw, params in (_TATONAS_PARAMETER_SEED.items() if isinstance(_TATONAS_PARAMETER_SEED, dict) else [])
    if isinstance(params, list)
}
TATONAS_SESSION_TTL = 15 * 60
TATONAS_STATION_CACHE_TTL = 6 * 60 * 60
# Sensor metadata changes far less frequently than telemetry values. Keeping it
# longer makes parameter selection effectively instant on warm Vercel instances.
TATONAS_SENSOR_CATALOG_TTL = 6 * 60 * 60
TATONAS_PARAMETER_CACHE_TTL = 6 * 60 * 60

TATONAS_STATION_PROFILE_BY_ID: dict[str, dict[str, str]] = {
    "4101": {"station_type": "rain_gauge", "logger_id": "9"},
    "4102": {"station_type": "rain_gauge", "logger_id": "9"},
    "4104": {"station_type": "rain_gauge", "logger_id": "8"},
    "4105": {"station_type": "water_level", "logger_id": "8"},
    "4106": {"station_type": "climatology", "logger_id": "9"},
}

# The plant-level sensor endpoint reports sensors per logger family (kd_logger),
# not per hardware id. These profiles keep known simple gauges from displaying
# climatology channels that are only physically installed at Sermo. Unknown/new
# stations fall back to their logger family + inferred station type below.
TATONAS_STATION_SENSOR_CODES: dict[str, list[str] | None] = {
    "4101": ["rainfall", "battery", "devicetemp"],
    "4102": ["rainfall", "battery", "devicetemp"],
    "4104": ["curahhujan", "battery", "devicetemp"],
    "4105": ["waterlevel", "debitair", "battery", "devicetemp"],
    # None = expose the complete logger-9 climatology catalog for Sermo.
    "4106": None,
}

TATONAS_SENSOR_LABELS = {
    "rainfall": "Curah Hujan",
    "rain": "Curah Hujan",
    "curah_hujan": "Curah Hujan",
    "precipitation": "Curah Hujan",
    "precipitation_intensity": "Intensitas Curah Hujan",
    "rainfall_intensity": "Intensitas Curah Hujan",
    "water_level": "Tinggi Muka Air",
    "waterlevel": "Tinggi Muka Air",
    "tma": "Tinggi Muka Air",
    "wind_speed": "Kecepatan Angin",
    "windspeed": "Kecepatan Angin",
    "battery": "Baterai Logger",
    "battery_logger": "Baterai Logger",
    "temperature": "Temperatur",
    "temp": "Temperatur",
    "humidity": "Kelembapan",
    "pressure": "Tekanan Udara",
    "solar_radiation": "Radiasi Matahari",
    "wind_direction": "Arah Angin",
    "airhumidity": "Kelembapan Udara",
    "airpressure": "Tekanan Udara",
    "airtemperatur": "Temperatur Udara",
    "devicetemp": "Temperatur Logger",
    "panlevel": "Pan Level",
    "radiation": "Radiasi Matahari",
    "uv": "Indeks UV",
    "winddirection": "Arah Angin",
    "windvelocity": "Kecepatan Angin",
    "curahhujan": "Curah Hujan",
    "debitair": "Debit",
}


def _tatonas_slug(value: Any) -> str:
    raw = clean_text(str(value or "")).lower()
    return re.sub(r"[^a-z0-9]+", "_", raw).strip("_")


def _tatonas_sensor_label(key: str, props: dict[str, Any]) -> str:
    candidates = [
        props.get("nm_sensor"), props.get("sensor_name"), props.get("name"),
        props.get("label"), props.get("kd_sensor"), key,
    ]
    # Try every exact alias first. This is important for upstream labels such as
    # "Wind Velocity" whose kd_sensor (windvelocity) has a better Indonesian
    # presentation name.
    for candidate in candidates:
        slug = _tatonas_slug(candidate)
        if slug in TATONAS_SENSOR_LABELS:
            return TATONAS_SENSOR_LABELS[slug]

    for candidate in candidates:
        slug = _tatonas_slug(candidate)
        if not slug:
            continue
        if "precipitation" in slug and "intens" in slug:
            return "Intensitas Curah Hujan"
        if "rain" in slug and "intens" in slug:
            return "Intensitas Curah Hujan"
        if slug in {"rainfall", "rain", "curah_hujan"} or ("curah" in slug and "hujan" in slug):
            return "Curah Hujan"
        if ("water" in slug and "level" in slug) or "tinggi_muka_air" in slug:
            return "Tinggi Muka Air"
        if (("wind" in slug and ("speed" in slug or "velocity" in slug)) or "kecepatan_angin" in slug):
            return "Kecepatan Angin"
        if "wind" in slug and "direction" in slug:
            return "Arah Angin"
        if "battery" in slug or "baterai" in slug:
            return "Baterai Logger"
        if "device" in slug and ("temp" in slug or "temperatur" in slug):
            return "Temperatur Logger"
        if "temperature" in slug or "temperatur" in slug or slug == "temp":
            return "Temperatur"
        if "humidity" in slug or "kelembapan" in slug:
            return "Kelembapan"
        if "pressure" in slug or "tekanan" in slug:
            return "Tekanan Udara"
        if "radiation" in slug or "radiasi" in slug:
            return "Radiasi Matahari"

    for candidate in candidates:
        if candidate:
            return clean_text(str(candidate)).replace("_", " ").title()
    return key.replace("_", " ").title()


def _tatonas_new_session() -> requests.Session:
    client = requests.Session()
    client.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/151.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    })
    return client


def _tatonas_csrf_from_html(text: str) -> str:
    soup = BeautifulSoup(text or "", "html.parser")
    meta = soup.find("meta", attrs={"name": "csrf-token"})
    if meta and meta.get("content"):
        return str(meta["content"])
    hidden = soup.find("input", attrs={"name": "_token"})
    if hidden and hidden.get("value"):
        return str(hidden["value"])
    return ""


def _tatonas_cookie_csrf(client: requests.Session) -> str:
    from urllib.parse import unquote
    for key, value in client.cookies.items():
        if key.upper().startswith("XSRF-TOKEN"):
            return unquote(value)
    return ""


def _tatonas_is_login_page(text: str, final_url: str) -> bool:
    if re.search(r'<input[^>]+type=["\']password["\']', text or "", re.I):
        return True
    return final_url.lower().rstrip("/").endswith("/admin/p/login")


def _tatonas_login(force: bool = False) -> requests.Session:
    global _TATONAS_SESSION, _TATONAS_SESSION_AT, _TATONAS_CSRF
    with TATONAS_LOCK:
        if (
            not force
            and _TATONAS_SESSION is not None
            and time.time() - _TATONAS_SESSION_AT < TATONAS_SESSION_TTL
        ):
            return _TATONAS_SESSION

        if (
            not TATONAS_USERNAME
            or not TATONAS_PASSWORD
            or TATONAS_USERNAME.startswith("ISI_")
            or TATONAS_PASSWORD.startswith("ISI_")
        ):
            raise RuntimeError(
                "TATONAS_USERNAME dan TATONAS_PASSWORD belum dikonfigurasi."
            )

        client = _tatonas_new_session()
        response = client.get(
            TATONAS_LOGIN_URL,
            headers={"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"},
            allow_redirects=True,
            timeout=TIMEOUT,
        )
        response.raise_for_status()

        csrf = _tatonas_csrf_from_html(response.text) or _tatonas_cookie_csrf(client)
        if not csrf:
            raise RuntimeError("Token CSRF login Tatonas tidak ditemukan.")

        payload = {
            "_token": csrf,
            "direct": "",
            "userid": TATONAS_USERNAME,
            "password": TATONAS_PASSWORD,
            "rememberme": "1",
        }
        login_response = client.post(
            TATONAS_LOGIN_URL,
            data=payload,
            headers={
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Content-Type": "application/x-www-form-urlencoded",
                "Origin": TATONAS_BASE_URL,
                "Referer": TATONAS_LOGIN_URL,
            },
            allow_redirects=False,
            timeout=TIMEOUT,
        )
        if login_response.status_code >= 400:
            login_response.raise_for_status()
        if login_response.status_code in (301, 302, 303, 307, 308):
            target = urljoin(TATONAS_LOGIN_URL, login_response.headers.get("Location", ""))
            if "/admin/p/home" not in target:
                raise RuntimeError("Login Tatonas ditolak atau redirect login tidak dikenali.")
        elif _tatonas_is_login_page(login_response.text, login_response.url):
            raise RuntimeError("Login Tatonas gagal. Periksa username/password.")

        verify = client.get(
            TATONAS_HOME_URL,
            headers={"Referer": TATONAS_LOGIN_URL},
            allow_redirects=True,
            timeout=TIMEOUT,
        )
        verify.raise_for_status()
        if _tatonas_is_login_page(verify.text, verify.url):
            raise RuntimeError("Login Tatonas gagal atau session tidak terbentuk.")

        _TATONAS_CSRF = _tatonas_csrf_from_html(verify.text) or csrf or _tatonas_cookie_csrf(client)
        _TATONAS_SESSION = client
        _TATONAS_SESSION_AT = time.time()
        return client


def _tatonas_clone_authenticated_session() -> requests.Session:
    master = _tatonas_login()
    client = _tatonas_new_session()
    client.cookies.update(master.cookies)
    return client


def _tatonas_ajax_headers(
    client: requests.Session,
    referer_text: str = "",
    *,
    json_accept: bool = False,
    referer_url: str = TATONAS_RAW_PAGE_URL,
) -> dict[str, str]:
    # Tatonas/Laravel exposes two different tokens in the HAR: X-CSRF-TOKEN
    # comes from the HTML meta tag, while X-XSRF-TOKEN comes from the XSRF
    # cookie. Do not mirror one value into both headers.
    csrf_token = _TATONAS_CSRF or _tatonas_csrf_from_html(referer_text)
    xsrf_token = _tatonas_cookie_csrf(client)
    headers = {
        "Accept": "application/json, text/plain, */*" if json_accept else "text/html, */*; q=0.01",
        "Referer": referer_url,
        "X-Requested-With": "XMLHttpRequest",
    }
    if csrf_token:
        headers["X-CSRF-TOKEN"] = csrf_token
    if xsrf_token:
        headers["X-XSRF-TOKEN"] = xsrf_token
    return headers


def _tatonas_infer_station_type(location: str) -> str:
    slug = _tatonas_slug(location)
    if any(token in slug for token in ("duga_air", "muka_air", "water_level", "waterlevel", "tma")):
        return "water_level"
    if any(token in slug for token in ("klimatologi", "climatology", "klimat")):
        return "climatology"
    return "rain_gauge"


def _tatonas_infer_logger_id(nm_logger: str, station_type: str) -> str:
    # The HAR sensor catalog associates rainfall/climatology channels with
    # logger family 9 (GPA JWT), and curahhujan/waterlevel with family 8 (GPA FTP).
    slug = _tatonas_slug(nm_logger)
    if "jwt" in slug:
        return "9"
    if "ftp" in slug:
        return "8"
    return "8" if station_type == "water_level" else "9"


def _tatonas_parse_hardware_html(text: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(text or "", "html.parser")
    result: list[dict[str, Any]] = []
    for tr in soup.select("tr.onclicktrlookup"):
        item: dict[str, Any] = {}
        raw = tr.get("data-json")
        if raw:
            try:
                item = json.loads(html.unescape(str(raw)))
            except Exception:
                item = {}
        if not item:
            cells = [td.get_text(" ", strip=True) for td in tr.find_all("td")]
            if len(cells) >= 5:
                item = {
                    "nm_logger": cells[0], "kd_hardware": cells[1], "location": cells[2],
                    "latitude": cells[3], "longitude": cells[4],
                }
        if not item:
            continue
        hw = clean_text(str(item.get("kd_hardware") or tr.get("data-name") or ""))
        if not hw:
            continue
        mapped = TATONAS_STATION_PROFILE_BY_ID.get(hw, {})
        item["kd_hardware"] = hw
        item["nm_logger"] = clean_text(str(item.get("nm_logger") or tr.get("data-id") or ""))
        item["location_original"] = clean_text(str(item.get("location") or ""))
        station_type = mapped.get("station_type") or _tatonas_infer_station_type(item["location_original"])
        item["name"] = station_alias("tatonas", hw, item["location_original"] or hw)
        item["station_type"] = station_type
        item["logger_id"] = mapped.get("logger_id") or _tatonas_infer_logger_id(item["nm_logger"], station_type)
        result.append(item)
    return result


def tatonas_stations(force: bool = False) -> list[dict[str, Any]]:
    global _TATONAS_STATION_CACHE
    with TATONAS_LOCK:
        if (
            not force and _TATONAS_STATION_CACHE
            and time.time() - _TATONAS_STATION_CACHE[0] < TATONAS_STATION_CACHE_TTL
        ):
            return _TATONAS_STATION_CACHE[1]

        client = _tatonas_login()
        page = client.get(TATONAS_RAW_PAGE_URL, timeout=TIMEOUT, allow_redirects=True)
        if _tatonas_is_login_page(page.text, page.url):
            client = _tatonas_login(force=True)
            page = client.get(TATONAS_RAW_PAGE_URL, timeout=TIMEOUT, allow_redirects=True)
        page.raise_for_status()
        response = client.get(
            TATONAS_HARDWARE_URL,
            params={"plant": TATONAS_PLANT, "logger": "undefined", "plantx": "undefined", "q": ""},
            headers=_tatonas_ajax_headers(client, page.text),
            timeout=TIMEOUT,
            allow_redirects=True,
        )
        if response.status_code in (401, 403) or _tatonas_is_login_page(response.text, response.url):
            client = _tatonas_login(force=True)
            page = client.get(TATONAS_RAW_PAGE_URL, timeout=TIMEOUT)
            response = client.get(
                TATONAS_HARDWARE_URL,
                params={"plant": TATONAS_PLANT, "logger": "undefined", "plantx": "undefined", "q": ""},
                headers=_tatonas_ajax_headers(client, page.text),
                timeout=TIMEOUT,
            )
        response.raise_for_status()
        rows = _tatonas_parse_hardware_html(response.text)
        # Keep all upstream hardware. The five current stations retain their
        # predefined short names, while newly added Tatonas stations are exposed
        # automatically using their upstream location and inferred station type.
        _TATONAS_STATION_CACHE = (time.time(), rows)
        _save_vendor_metadata("tatonas", "positions.json", rows)
        return rows


def _tatonas_station(hw: str) -> dict[str, Any]:
    for row in tatonas_stations():
        if str(row.get("kd_hardware")) == hw:
            return row
    for row in tatonas_stations(force=True):
        if str(row.get("kd_hardware")) == hw:
            return row
    if hw in TATONAS_STATION_PROFILE_BY_ID:
        mapped = TATONAS_STATION_PROFILE_BY_ID[hw]
        return {
            "kd_hardware": hw,
            "nm_logger": "",
            "location_original": hw,
            "name": station_alias("tatonas", hw, hw),
            **mapped,
        }
    raise RuntimeError(f"Hardware Tatonas {hw} tidak ditemukan.")


def _tatonas_sensor_catalog(force: bool = False) -> list[dict[str, Any]]:
    global _TATONAS_SENSOR_CATALOG_CACHE
    with TATONAS_LOCK:
        if (
            not force
            and _TATONAS_SENSOR_CATALOG_CACHE
            and time.time() - _TATONAS_SENSOR_CATALOG_CACHE[0] < TATONAS_SENSOR_CATALOG_TTL
        ):
            return _TATONAS_SENSOR_CATALOG_CACHE[1]

        client = _tatonas_login()
        page = client.get(TATONAS_HOME_URL, timeout=TIMEOUT, allow_redirects=True)
        if _tatonas_is_login_page(page.text, page.url):
            client = _tatonas_login(force=True)
            page = client.get(TATONAS_HOME_URL, timeout=TIMEOUT, allow_redirects=True)
        page.raise_for_status()

        response = client.get(
            TATONAS_SENSOR_CATALOG_URL,
            params={"plant": TATONAS_PLANT},
            headers=_tatonas_ajax_headers(client, page.text, json_accept=True, referer_url=TATONAS_HOME_URL),
            timeout=TIMEOUT,
            allow_redirects=True,
        )
        if response.status_code in (401, 403) or _tatonas_is_login_page(response.text, response.url):
            client = _tatonas_login(force=True)
            response = client.get(
                TATONAS_SENSOR_CATALOG_URL,
                params={"plant": TATONAS_PLANT},
                headers=_tatonas_ajax_headers(client, json_accept=True, referer_url=TATONAS_HOME_URL),
                timeout=TIMEOUT,
                allow_redirects=True,
            )
        response.raise_for_status()
        try:
            payload = response.json()
        except ValueError as exc:
            raise RuntimeError("Response katalog sensor Tatonas bukan JSON.") from exc

        rows = payload.get("data_sensor", []) if isinstance(payload, dict) else []
        catalog: list[dict[str, Any]] = []
        for raw in rows:
            if not isinstance(raw, dict):
                continue
            code = clean_text(str(raw.get("kd_sensor") or ""))
            if not code:
                continue
            props = {
                "kd_sensor": code,
                "nm_sensor": raw.get("nm_sensor"),
                "satuan": raw.get("satuan"),
            }
            catalog.append({
                "id": code,
                "sensor_code": code,
                "name": _tatonas_sensor_label(code, props),
                "unit": clean_text(str(raw.get("satuan") or "")),
                "logger_id": clean_text(str(raw.get("kd_logger") or "")),
                "sensor_type": clean_text(str(raw.get("type") or "")),
            })
        if not catalog:
            raise RuntimeError("Katalog sensor Tatonas kosong.")
        _TATONAS_SENSOR_CATALOG_CACHE = (time.time(), catalog)
        return catalog


def _tatonas_station_sensor_codes(station: dict[str, Any]) -> list[str] | None:
    hw = clean_text(str(station.get("kd_hardware") or ""))
    if hw in TATONAS_STATION_SENSOR_CODES:
        return TATONAS_STATION_SENSOR_CODES[hw]
    station_type = station.get("station_type") or "rain_gauge"
    logger_id = clean_text(str(station.get("logger_id") or ""))
    if station_type == "climatology":
        return None
    if station_type == "water_level":
        return ["waterlevel", "debitair", "battery", "devicetemp"]
    return ["curahhujan", "battery", "devicetemp"] if logger_id == "8" else ["rainfall", "battery", "devicetemp"]


def _tatonas_parameters_from_catalog(hw: str, catalog: list[dict[str, Any]], station: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    station = station or _tatonas_station(hw)
    logger_id = clean_text(str(station.get("logger_id") or "")) or _tatonas_infer_logger_id(
        clean_text(str(station.get("nm_logger") or "")),
        clean_text(str(station.get("station_type") or "rain_gauge")),
    )
    allowed_codes = _tatonas_station_sensor_codes(station)
    station_type = clean_text(str(station.get("station_type") or "rain_gauge"))

    selected = [item for item in catalog if clean_text(str(item.get("logger_id") or "")) == logger_id]
    if station_type == "rain_gauge":
        # Rain channels + logger health channels. This automatically includes
        # future Rain sensors (e.g. a precipitation-intensity channel) without
        # exposing Sermo-only climatology channels on a simple rain gauge.
        selected = [item for item in selected if str(item.get("sensor_type")) in {"Rain", "Device"}]
    elif station_type == "water_level":
        selected = [item for item in selected if str(item.get("sensor_type")) in {"Water", "Factory", "Device"}]

    order = {code: idx for idx, code in enumerate(allowed_codes or [])}
    if station_type == "climatology":
        # Sermo: put rainfall first, then climatology, then device channels.
        type_order = {"Rain": 0, "Climatology": 1, "Factory": 2, "Device": 3}
        selected.sort(key=lambda x: (type_order.get(str(x.get("sensor_type")), 9), str(x.get("name", ""))))
    else:
        selected.sort(key=lambda x: (order.get(str(x.get("sensor_code")), 50), str(x.get("name", ""))))

    params: list[dict[str, Any]] = []
    for sensor in selected:
        name = clean_text(str(sensor.get("name") or sensor.get("sensor_code") or "Sensor"))
        unit = clean_text(str(sensor.get("unit") or ""))
        code = clean_text(str(sensor.get("sensor_code") or sensor.get("id") or ""))
        is_tma = bool(re.search(r"water\s*level|tinggi\s*muka\s*air|\btma\b|muka\s*air", name, re.I)) or code == "waterlevel"
        is_rain = bool(re.search(r"rain|curah\s*hujan|precip", name, re.I)) or code in {"rainfall", "curahhujan"}
        params.append({
            "id": code,
            "name": "Tinggi Muka Air" if is_tma else name,
            "type": "tma" if is_tma else ("rain" if is_rain else "sensor"),
            "unit": "m" if is_tma else unit,
            "source_unit": "cm" if is_tma else unit,
            "sensor_code": code,
            "sensor_type": sensor.get("sensor_type", ""),
        })
    return params


def _tatonas_find_timestamp_list(payload: Any, target_len: int) -> list[Any]:
    priority = {
        "timestamp", "timestamps", "datetime", "datetimes", "waktu", "tanggal",
        "labels", "label", "categories", "category",
    }
    found: list[list[Any]] = []

    def walk(obj: Any, depth: int = 0) -> None:
        if depth > 6:
            return
        if isinstance(obj, dict):
            for key, value in obj.items():
                if _tatonas_slug(key) in priority and isinstance(value, list) and len(value) == target_len:
                    found.append(value)
                walk(value, depth + 1)
        elif isinstance(obj, list):
            for value in obj[:20]:
                walk(value, depth + 1)

    walk(payload)
    return found[0] if found else []


def _tatonas_generated_timestamps(t1: str, count: int, vq: str) -> list[str]:
    dt = _parse_any_datetime(t1)
    if dt is None:
        return [str(i + 1) for i in range(count)]
    step = timedelta(days=1) if vq.lower() == "perhari" else timedelta(hours=1)
    return [(dt + i * step).strftime("%Y-%m-%d %H:%M:%S") for i in range(count)]


def _tatonas_normalize_sensor_payload(payload: Any, t1: str, vq: str) -> tuple[list[dict[str, Any]], list[Any]]:
    """Normalize Tatonas response into sensor series + aligned timestamps.

    Current Tatonas responses expose the most reliable telemetry representation
    in ``data_table``.  Each row has one ``date_act`` and a ``sensor`` mapping,
    whose child entries contain ``value`` + ``properties``.  Older code looked
    only for ``nilai`` arrays near the root and therefore missed the actual
    ``data_graph.datax.graph`` / ``data_graph.graph.data`` nesting.

    Prefer data_table because it preserves gaps naturally (e.g. Sempor may only
    return 07:00-17:00), then fall back to graph-style arrays for compatibility.
    """
    sensors_by_code: dict[str, dict[str, Any]] = {}
    timestamps: list[Any] = []

    # --- Primary path: explicit table rows ---------------------------------
    table = payload.get("data_table") if isinstance(payload, dict) else None
    if isinstance(table, list):
        for row in table:
            if not isinstance(row, dict):
                continue
            stamp = row.get("date_act") or row.get("datetime") or row.get("timestamp") or row.get("waktu")
            sensor_map = row.get("sensor")
            if not isinstance(sensor_map, dict) or stamp in (None, ""):
                continue
            row_index = len(timestamps)
            timestamps.append(stamp)
            # Keep every existing series aligned with this new timestamp.
            for series in sensors_by_code.values():
                series["values"].append(None)

            for key, sensor_obj in sensor_map.items():
                if not isinstance(sensor_obj, dict):
                    continue
                props = sensor_obj.get("properties") if isinstance(sensor_obj.get("properties"), dict) else {}
                code = clean_text(str(props.get("kd_sensor") or key)) or clean_text(str(key))
                if not code:
                    continue
                if code not in sensors_by_code:
                    sensors_by_code[code] = {
                        "id": code,
                        "sensor_code": code,
                        "name": _tatonas_sensor_label(code, props),
                        "unit": clean_text(str(props.get("unit") or props.get("satuan") or props.get("unit_sensor") or "")),
                        "values": [None] * (row_index + 1),
                        "properties": props,
                    }
                series = sensors_by_code[code]
                # Metadata can be more complete on a later row.
                if props:
                    series["properties"] = props
                    series["name"] = _tatonas_sensor_label(code, props)
                    series["unit"] = clean_text(str(props.get("unit") or props.get("satuan") or props.get("unit_sensor") or series.get("unit") or ""))
                value = sensor_obj.get("value")
                if value is None and not isinstance(sensor_obj.get("nilai"), list):
                    value = sensor_obj.get("nilai")
                series["values"][row_index] = value

        if sensors_by_code and timestamps:
            return list(sensors_by_code.values()), timestamps

    # --- Compatibility path: graph arrays ----------------------------------
    sensors: list[dict[str, Any]] = []
    seen_codes: set[str] = set()

    def add_sensor(key: Any, obj: Any) -> None:
        if not isinstance(obj, dict):
            return
        values = obj.get("nilai")
        if not isinstance(values, list):
            for alt in ("values", "data", "value"):
                if isinstance(obj.get(alt), list):
                    values = obj[alt]
                    break
        if not isinstance(values, list):
            return
        props = obj.get("properties") if isinstance(obj.get("properties"), dict) else {}
        sensor_code = clean_text(str(props.get("kd_sensor") or key))
        code_key = sensor_code or clean_text(str(key))
        if not code_key or code_key in seen_codes:
            return
        seen_codes.add(code_key)
        sensors.append({
            "id": code_key,
            "sensor_code": sensor_code or code_key,
            "name": _tatonas_sensor_label(sensor_code or str(key), props),
            "unit": clean_text(str(props.get("unit") or props.get("satuan") or props.get("unit_sensor") or "")),
            "values": values,
            "properties": props,
        })

    def walk_graph(obj: Any, depth: int = 0) -> None:
        if depth > 8:
            return
        if isinstance(obj, dict):
            for key, value in obj.items():
                add_sensor(key, value)
                if isinstance(value, (dict, list)):
                    walk_graph(value, depth + 1)
        elif isinstance(obj, list):
            for value in obj[:50]:
                if isinstance(value, (dict, list)):
                    walk_graph(value, depth + 1)

    walk_graph(payload)
    max_len = max((len(sensor["values"]) for sensor in sensors), default=0)
    timestamps = _tatonas_find_timestamp_list(payload, max_len) if max_len else []

    # Tatonas graph labels are commonly a dictionary keyed by datetime rather
    # than a list, e.g. data_graph.graph.label.{"2026-05-15 00:00": {...}}.
    if max_len and not timestamps and isinstance(payload, dict):
        label_obj = (((payload.get("data_graph") or {}).get("graph") or {}).get("label")
                     if isinstance(payload.get("data_graph"), dict) else None)
        if isinstance(label_obj, dict) and len(label_obj) == max_len:
            timestamps = [
                (v.get("date") if isinstance(v, dict) and v.get("date") else k)
                for k, v in label_obj.items()
            ]

    if max_len and not timestamps:
        timestamps = _tatonas_generated_timestamps(t1, max_len, vq)
    return sensors, timestamps

def _tatonas_fetch_payload(
    hw: str,
    dari: str,
    sampai: str,
    vq: str = "Perjam",
    *,
    isolated_session: bool = False,
) -> tuple[Any, list[dict[str, Any]], list[Any]]:
    client = _tatonas_clone_authenticated_session() if isolated_session else _tatonas_login()
    page = client.get(TATONAS_RAW_PAGE_URL, timeout=TIMEOUT, allow_redirects=True)
    if _tatonas_is_login_page(page.text, page.url):
        _tatonas_login(force=True)
        client = _tatonas_clone_authenticated_session() if isolated_session else _tatonas_login()
        page = client.get(TATONAS_RAW_PAGE_URL, timeout=TIMEOUT, allow_redirects=True)
    page.raise_for_status()

    safe_from, safe_to = _tatonas_clamp_range(dari, sampai)
    params = {
        "hw": hw,
        "plant": TATONAS_PLANT,
        "t1": _tatonas_format_datetime(safe_from),
        "t2": _tatonas_format_datetime(safe_to),
        "vq": vq,
    }
    response = client.get(
        TATONAS_DATA_URL,
        params=params,
        headers=_tatonas_ajax_headers(client, page.text, json_accept=True),
        timeout=max(TIMEOUT, 60),
        allow_redirects=True,
    )
    if response.status_code in (401, 403) or _tatonas_is_login_page(response.text, response.url):
        _tatonas_login(force=True)
        client = _tatonas_clone_authenticated_session() if isolated_session else _tatonas_login()
        page = client.get(TATONAS_RAW_PAGE_URL, timeout=TIMEOUT)
        response = client.get(
            TATONAS_DATA_URL,
            params=params,
            headers=_tatonas_ajax_headers(client, page.text, json_accept=True),
            timeout=max(TIMEOUT, 60),
        )
    response.raise_for_status()
    try:
        payload = response.json()
    except ValueError as exc:
        raise RuntimeError("Response data Tatonas bukan JSON.") from exc
    sensors, timestamps = _tatonas_normalize_sensor_payload(payload, params["t1"], vq)
    return payload, sensors, timestamps


def _tatonas_format_datetime(value: str) -> str:
    dt = _parse_any_datetime(value)
    if dt is None:
        raw = clean_text(value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}", raw):
            return raw + ":00"
        return raw
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _tatonas_clamp_range(dari: str, sampai: str) -> tuple[str, str]:
    """Clamp outbound Tatonas requests so t2 never points into the future.

    The UI may represent a whole current month/year, while the upstream Tatonas
    endpoint is happiest when the requested end time is not later than now.
    """
    start = _parse_any_datetime(dari)
    end = _parse_any_datetime(sampai)
    now = datetime.now()
    if start is None or end is None:
        return dari, sampai
    if start > now:
        raise RuntimeError("Periode Tatonas dimulai di masa depan dan belum memiliki data.")
    if end > now:
        end = now
    if end < start:
        end = start
    return start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")


def _tatonas_parameter_probe_range(dari: str = "", sampai: str = "") -> tuple[str, str]:
    """Return a small <=24 h range used only to discover sensor names.

    Parameter discovery must stay cheap on Vercel and must not download a whole
    month/year merely to populate a dropdown.
    """
    now = datetime.now()
    start = _parse_any_datetime(dari) if dari else None
    end = _parse_any_datetime(sampai) if sampai else None

    if start is None:
        start = now - timedelta(hours=23)
    if start > now:
        start = now - timedelta(hours=23)

    probe_end = start + timedelta(hours=23, minutes=59, seconds=59)
    if end is not None:
        probe_end = min(probe_end, end)
    probe_end = min(probe_end, now)
    if probe_end < start:
        start = max(now - timedelta(hours=23), datetime(2000, 1, 1))
        probe_end = now

    return start.strftime("%Y-%m-%d %H:%M:%S"), probe_end.strftime("%Y-%m-%d %H:%M:%S")


def tatonas_parameters(hw: str, dari: str = "", sampai: str = "", force: bool = False) -> list[dict[str, Any]]:
    cached = _TATONAS_PARAMETER_CACHE.get(hw)
    if not force and cached and time.time() - cached[0] < TATONAS_PARAMETER_CACHE_TTL:
        return cached[1]

    # Fast path: discover channels from Tatonas' plant-level sensor metadata.
    # This does not depend on whether a particular date contains telemetry data.
    catalog_error: Exception | None = None
    try:
        catalog = _tatonas_sensor_catalog(force=force)
        params = _tatonas_parameters_from_catalog(hw, catalog)
        if params:
            _TATONAS_PARAMETER_CACHE[hw] = (time.time(), params)
            _save_vendor_metadata(
                "tatonas", "parameter_catalog.json",
                {key: value[1] for key, value in _TATONAS_PARAMETER_CACHE.items()},
            )
            return params
    except Exception as exc:
        catalog_error = exc

    # Compatibility fallback if the upstream metadata endpoint changes. Probe at
    # most 24 h, matching the old v3.1 behaviour, but never use this as primary.
    _tatonas_station(hw)
    probe_from, probe_to = _tatonas_parameter_probe_range(dari, sampai)
    try:
        _, sensors, _ = _tatonas_fetch_payload(hw, probe_from, probe_to, "Perjam")
    except Exception:
        if catalog_error:
            raise RuntimeError(f"Katalog sensor Tatonas gagal dimuat: {catalog_error}") from catalog_error
        raise

    params: list[dict[str, Any]] = []
    for sensor in sensors:
        name = sensor["name"]
        unit = sensor.get("unit", "")
        code = clean_text(str(sensor.get("sensor_code") or sensor.get("id") or ""))
        is_tma = bool(re.search(r"water\s*level|tinggi\s*muka\s*air|\btma\b|muka\s*air", name, re.I)) or code == "waterlevel"
        params.append({
            "id": code or sensor["id"],
            "name": "Tinggi Muka Air" if is_tma else name,
            "type": "tma" if is_tma else ("rain" if re.search(r"rain|curah\s*hujan|precip", name, re.I) else "sensor"),
            "unit": "m" if is_tma else unit,
            "source_unit": "cm" if is_tma else unit,
            "sensor_code": code,
        })
    if not params:
        if catalog_error:
            raise RuntimeError(f"Sensor Tatonas tidak ditemukan; katalog metadata juga gagal: {catalog_error}") from catalog_error
        raise RuntimeError("Sensor Tatonas tidak ditemukan pada data.")
    _TATONAS_PARAMETER_CACHE[hw] = (time.time(), params)
    _save_vendor_metadata(
        "tatonas", "parameter_catalog.json",
        {key: value[1] for key, value in _TATONAS_PARAMETER_CACHE.items()},
    )
    return params


def _tatonas_data_single(
    hw: str,
    dari: str,
    sampai: str,
    sensor_id: str,
    *,
    isolated_session: bool = False,
) -> tuple[list[str], list[list[Any]], dict[str, Any], dict[str, Any]]:
    station = _tatonas_station(hw)
    _, sensors, timestamps = _tatonas_fetch_payload(
        hw, dari, sampai, "Perjam", isolated_session=isolated_session
    )
    requested = _tatonas_slug(sensor_id)
    selected = next(
        (
            sensor for sensor in sensors
            if requested in {
                _tatonas_slug(sensor.get("id")),
                _tatonas_slug(sensor.get("sensor_code")),
            }
        ),
        None,
    )
    if selected is None and requested in {"rainfall", "curahhujan", "curah_hujan", "rain", "precipitation"}:
        # Some Tatonas logger families name the physical rain channel
        # ``rainfall`` while others use ``curahhujan``. Treat these as the same
        # primary rainfall parameter when metadata and telemetry use different
        # aliases.
        selected = next(
            (
                sensor for sensor in sensors
                if _tatonas_slug((sensor.get("properties") or {}).get("kd_type")) == "rain"
                or _tatonas_slug(sensor.get("sensor_code")) in {"rainfall", "curahhujan"}
            ),
            None,
        )
    if selected is None:
        # Sensor validity comes from metadata, not from one telemetry slice. A
        # valid channel may be absent in an early/late sparse chunk while other
        # channels are still present. Treat that slice as empty instead of error.
        meta = next(
            (p for p in tatonas_parameters(hw) if _tatonas_slug(p.get("id")) == requested),
            None,
        )
        if meta is not None or not sensors:
            meta = meta or {"id": sensor_id, "name": sensor_id, "type": "sensor", "unit": ""}
            return ["Waktu", str(meta.get("name") or sensor_id)], [], station, meta
        available = ", ".join(clean_text(str(sensor.get("sensor_code") or sensor.get("id") or "")) for sensor in sensors)
        suffix = f" Sensor tersedia: {available}." if available else ""
        raise RuntimeError("Parameter Tatonas yang dipilih tidak ditemukan periode tersebut." + suffix)

    label = selected["name"]
    values = selected.get("values", [])
    count = min(len(timestamps), len(values))
    rows: list[list[Any]] = []
    for i in range(count):
        value = values[i]
        if value is None or clean_text(str(value)) == "":
            continue
        try:
            number: Any = float(value)
        except (TypeError, ValueError):
            number = value
        dt = _parse_any_datetime(timestamps[i])
        stamp = dt.strftime("%Y-%m-%d %H:%M:%S") if dt else clean_text(str(timestamps[i]))
        rows.append([stamp, number])

    is_tma = station.get("station_type") == "water_level" and bool(re.search(r"water\s*level|tinggi\s*muka\s*air|\btma\b|muka\s*air", label, re.I))
    parameter = {
        "id": selected["id"],
        "name": "Tinggi Muka Air" if is_tma else label,
        "type": "tma" if is_tma else ("rain" if re.search(r"rain|curah\s*hujan|precip", label, re.I) else "sensor"),
        # Keep raw cm here: the repo's existing Tatonas processor converts TMA cm -> m exactly once.
        "unit": "m" if is_tma else selected.get("unit", ""),
        "source_unit": "cm" if is_tma else selected.get("unit", ""),
    }
    return ["Waktu", parameter["name"]], rows, station, parameter


def tatonas_data(
    hw: str,
    dari: str,
    sampai: str,
    sensor_id: str,
    *,
    isolated_session: bool = False,
) -> tuple[list[str], list[list[Any]], dict[str, Any], dict[str, Any]]:
    """Fetch Tatonas in <= configured calendar-month chunks with timeout split."""
    start = _parse_any_datetime(dari)
    end = _parse_any_datetime(sampai)
    if start is None or end is None:
        raise RuntimeError("Format periode Tatonas tidak valid.")
    if end < start:
        raise RuntimeError("Tanggal akhir tidak boleh sebelum tanggal awal.")

    chunks = _split_datetime_month_chunks(start, end, TATONAS_CHUNK_MONTHS)
    station = _tatonas_station(hw)
    headers: list[str] = []
    parameter: dict[str, Any] = next(
        (p for p in tatonas_parameters(hw) if _tatonas_slug(p.get("id")) == _tatonas_slug(sensor_id)),
        {"id": sensor_id, "name": sensor_id, "type": "sensor", "unit": ""},
    )
    merged: list[list[Any]] = []

    def fetch_resilient(a: datetime, b: datetime, depth: int = 0) -> tuple[list[str], list[list[Any]], dict[str, Any]]:
        last_exc: Exception | None = None
        for attempt in range(2):
            try:
                h, r, _st, prm = _tatonas_data_single(
                    hw, a.strftime("%Y-%m-%d %H:%M"), b.strftime("%Y-%m-%d %H:%M"),
                    sensor_id, isolated_session=isolated_session
                )
                return h, r, prm
            except Exception as exc:
                last_exc = exc
                if attempt == 0:
                    time.sleep(0.15)
        assert last_exc is not None
        span = b - a
        if _timeout_like(last_exc) and depth < 3 and span > timedelta(days=2):
            midpoint = a + span / 2
            midpoint = midpoint.replace(second=0, microsecond=0)
            left_h, left_r, left_p = fetch_resilient(a, midpoint, depth + 1)
            right_start = midpoint + timedelta(minutes=1)
            right_h, right_r, right_p = fetch_resilient(right_start, b, depth + 1)
            return left_h or right_h, left_r + right_r, right_p or left_p
        raise last_exc

    for a, b in chunks:
        h, rows, prm = fetch_resilient(a, b)
        if h and not headers:
            headers = h
        if prm:
            parameter = prm
        merged.extend(rows)

    seen: set[tuple[str, str]] = set()
    rows_out: list[list[Any]] = []
    for row in sorted(merged, key=lambda r: str(r[0]) if r else ""):
        if not row:
            continue
        key = (str(row[0]), str(row[1] if len(row) > 1 else ""))
        if key in seen:
            continue
        seen.add(key)
        rows_out.append(row)
    return headers or ["Waktu", str(parameter.get("name") or sensor_id)], rows_out, station, parameter


# ============================================================
# PARAMETER CACHE
# ============================================================

PARAMETER_CACHE: dict[
    str,
    tuple[float, list[dict[str, str]]]
] = {}

POSITION_TYPE_CACHE: dict[
    str,
    tuple[float, list[dict[str, str]]]
] = {}


# ============================================================
# PARAMETER CATALOG
# ============================================================

PARAMETER_CATALOG_SOURCE_PATH = (
    ROOT_DIR
    / "data"
    / "beacon"
    / "parameter_catalog.json"
)

if os.environ.get("VERCEL"):
    PARAMETER_CATALOG_PATH = (
        Path("/tmp")
        / "bbws_parameter_catalog.json"
    )
else:
    PARAMETER_CATALOG_PATH = (
        PARAMETER_CATALOG_SOURCE_PATH
    )

CATALOG_LOCK = threading.Lock()

CATALOG_WARMING: set[str] = set()


# ============================================================
# GLOBAL BBWS SESSION
# ============================================================
#
# Sangat penting untuk Vercel.
#
# Jangan membuat:
#
#     BBWSSession()
#
# setiap request jika tidak diperlukan.
#
# Instance Vercel biasanya dapat menangani beberapa request
# selama warm container masih hidup.
#
# Dengan global session:
#
#     LOGIN
#       ↓
#     requests.Session()
#       ↓
#     cookie tetap
#       ↓
#     request berikutnya dapat reuse session
#
# ============================================================

BBWS_CLIENT_LOCK = threading.RLock()

_GLOBAL_BBWS_CLIENT: "BBWSSession | None" = None


def get_bbws_client() -> "BBWSSession":
    global _GLOBAL_BBWS_CLIENT

    with BBWS_CLIENT_LOCK:

        if _GLOBAL_BBWS_CLIENT is None:
            _GLOBAL_BBWS_CLIENT = BBWSSession()

        return _GLOBAL_BBWS_CLIENT


def reset_bbws_client() -> "BBWSSession":
    global _GLOBAL_BBWS_CLIENT

    with BBWS_CLIENT_LOCK:

        _GLOBAL_BBWS_CLIENT = BBWSSession()

        return _GLOBAL_BBWS_CLIENT


# ============================================================
# EMPTY DATA EXCEPTION
# ============================================================

class EmptyHistoricalData(Exception):
    """
    Data memang kosong.

    Ini berbeda dengan:
    - 520
    - 522
    - timeout
    - connection error
    - 500
    - 502
    - 503
    """

    pass


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_text(value: str) -> str:
    return re.sub(
        r"\s+",
        " ",
        value or "",
    ).strip()


def display_position_name(name: str) -> str:
    parts = clean_text(name).split()
    displayed = " ".join(parts[2:]) if len(parts) > 2 else clean_text(name)
    # Beacon labels use "Bendung" as a type prefix; hide it from the operator
    # name, but keep the genuine place name "Bendungan" unchanged.
    displayed = re.sub(r"^Bendung\s+", "", displayed, flags=re.I)
    return clean_text(displayed)


def parse_html(html: str) -> BeautifulSoup:
    return BeautifulSoup(
        html or "",
        "html.parser",
    )


def looks_like_login(
    html: str,
    url: str = "",
) -> bool:

    path = (
        urlparse(url).path.lower()
        if url
        else ""
    )

    if path == "/login":
        return True

    if path.startswith("/login/"):
        return True

    soup = parse_html(html)

    return bool(
        soup.find(
            "input",
            attrs={"type": "password"},
        )
        and soup.find("form")
    )


# ============================================================
# TOKEN
# ============================================================

def extract_token(
    html: str,
    url: str | None = None,
) -> str | None:

    soup = parse_html(html)

    node = soup.find(
        "input",
        attrs={"name": "token"},
    )

    if node and node.get("value"):
        return node.get("value").strip()

    patterns = [
        r'name=["\']token["\'][^>]*value=["\']([^"\']+)',
        r'value=["\']([^"\']+)["\'][^>]*name=["\']token["\']',
        r'["\']token["\']\s*[:=]\s*["\']([^"\']+)',
    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            html or "",
            flags=re.I,
        )

        if match:
            return match.group(1).strip()

    if url:

        match = re.search(
            r"/analisa/data/([^/?#]+)",
            urlparse(url).path,
        )

        if match:
            return match.group(1)

    return None


# ============================================================
# SELECT OPTIONS
# ============================================================

def extract_options(
    html: str,
    select_id: str,
) -> list[dict[str, str]]:

    soup = parse_html(html)

    select = soup.find(
        "select",
        id=select_id,
    )

    if not select:
        return []

    out: list[dict[str, str]] = []

    for option in select.find_all("option"):

        value = clean_text(
            option.get("value", "")
        )

        label = clean_text(
            option.get_text(
                " ",
                strip=True,
            )
        )

        if value:

            out.append(
                {
                    "id": value,
                    "name": label,
                }
            )

    return out


# ============================================================
# PARAMETER CLASSIFICATION
# ============================================================

def classify_parameter(
    name: str,
) -> str:

    s = clean_text(name).lower()

    if any(
        k in s
        for k in [
            "debit",
            "discharge",
            "flow rate",
            "flowrate",
            "streamflow",
        ]
    ):
        return "discharge"

    if (
        "tinggi muka air" in s
        or "water level" in s
        or "waterlevel" in s
        or "water stage" in s
        or "stage" in s
        or "elevasi muka air" in s
        or "muka air" in s
        or re.search(
            r"\belv\.?\s*ma\b",
            s,
        )
    ):
        return "water_level"

    if any(
        k in s
        for k in [
            "curah hujan",
            "precipitation intensity",
            "precipitation",
            "rainfall",
            "rain intensity",
        ]
    ):
        return "rainfall"

    if any(
        k in s
        for k in [
            "battery",
            "baterai",
        ]
    ):
        return "battery"

    if any(
        k in s
        for k in [
            "temperatur",
            "temperature",
        ]
    ):
        return "temperature"

    return "other"


def parameter_with_metadata(
    items: list[dict[str, str]],
) -> list[dict[str, str]]:

    result = []

    for item in items:

        name = item.get(
            "name",
            "",
        )

        result.append(
            {
                "id": item.get("id", ""),
                "name": name,
                "type": classify_parameter(name),
            }
        )

    return result


def parameter_candidates_for_data_type(
    items: list[dict[str, str]],
    data_type: str,
) -> list[dict[str, str]]:

    target = (
        "rainfall"
        if data_type == "rain"
        else "water_level"
        if data_type == "tma"
        else None
    )

    if target is None:
        return parameter_with_metadata(items)

    return [
        x
        for x in parameter_with_metadata(items)
        if x["type"] == target
    ]


# ============================================================
# LOAD PARAMETER CATALOG
# ============================================================

def _load_parameter_catalog() -> None:

    catalog_path: Path | None = None

    # Prioritas:
    # 1. runtime cache
    # 2. repository

    if PARAMETER_CATALOG_PATH.exists():

        catalog_path = PARAMETER_CATALOG_PATH

    elif PARAMETER_CATALOG_SOURCE_PATH.exists():

        catalog_path = PARAMETER_CATALOG_SOURCE_PATH

    if catalog_path is None:
        print(
            "Parameter catalog tidak ditemukan."
        )
        return

    try:

        with catalog_path.open(
            "r",
            encoding="utf-8",
        ) as f:

            raw = json.load(f) or {}

        if not isinstance(raw, dict):
            return

        loaded = 0

        for logger_id, items in raw.items():

            if not isinstance(items, list):
                continue

            normalized = (
                parameter_with_metadata(items)
            )

            PARAMETER_CACHE[
                str(logger_id)
            ] = (
                time.time(),
                normalized,
            )

            loaded += 1

        print(
            "Parameter catalog loaded: "
            f"{loaded} posisi dari {catalog_path}"
        )

    except (
        OSError,
        json.JSONDecodeError,
    ) as exc:

        print(
            f"Gagal membaca parameter catalog: {exc}"
        )


_load_parameter_catalog()


# ============================================================
# SAVE PARAMETER CATALOG
# ============================================================

def _save_parameter_catalog() -> None:

    try:

        raw = {
            key: value[1]
            for key, value
            in PARAMETER_CACHE.items()
        }

        PARAMETER_CATALOG_PATH.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        tmp = PARAMETER_CATALOG_PATH.with_suffix(
            ".tmp"
        )

        with tmp.open(
            "w",
            encoding="utf-8",
        ) as f:

            json.dump(
                raw,
                f,
                ensure_ascii=False,
                indent=2,
            )

        tmp.replace(
            PARAMETER_CATALOG_PATH
        )

    except Exception as exc:

        print(
            f"Gagal menyimpan parameter catalog: {exc}"
        )


# ============================================================
# LOGIN FORM
# ============================================================

def find_login_form(
    html: str,
) -> tuple[
    str,
    dict[str, str],
    str,
    str,
]:

    soup = parse_html(html)

    forms = soup.find_all("form")

    if not forms:

        raise RuntimeError(
            "Form login tidak ditemukan pada halaman sumber."
        )

    chosen = None

    for form in forms:

        if form.find(
            "input",
            attrs={"type": "password"},
        ):

            chosen = form
            break

    if chosen is None:
        chosen = forms[0]

    action = urljoin(
        BASE_URL + "/",
        chosen.get("action") or "/login",
    )

    payload: dict[str, str] = {}

    for inp in chosen.find_all("input"):

        name = clean_text(
            inp.get("name", "")
        )

        if name:

            payload[name] = inp.get(
                "value",
                "",
            )

    password_input = chosen.find(
        "input",
        attrs={"type": "password"},
    )

    if not password_input:

        raise RuntimeError(
            "Input password tidak ditemukan pada form login."
        )

    password_name = (
        password_input.get("name")
        or PASSWORD_FIELD
    )

    username_input = chosen.find(
        "input",
        attrs={"name": USERNAME_FIELD},
    )

    if username_input is None:

        for candidate in [
            "username",
            "user",
            "nama_pengguna",
            "email",
            "login",
        ]:

            username_input = chosen.find(
                "input",
                attrs={"name": candidate},
            )

            if username_input:
                break

    if username_input is None:

        for inp in chosen.find_all("input"):

            typ = (
                inp.get("type")
                or "text"
            ).lower()

            name = inp.get("name") or ""

            if typ in {
                "text",
                "email",
            } and name:

                username_input = inp
                break

    if username_input is None:

        raise RuntimeError(
            "Input username tidak ditemukan pada form login."
        )

    username_name = (
        username_input.get("name")
        or USERNAME_FIELD
    )

    return (
        action,
        payload,
        username_name,
        password_name,
    )


# ============================================================
# BBWS SESSION
# ============================================================

class BBWSSession:

    def __init__(self) -> None:

        self.session = requests.Session()

        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 "
                    "(Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 "
                    "Chrome/151.0 Safari/537.36"
                ),
                "Accept": (
                    "text/html,"
                    "application/xhtml+xml,"
                    "application/xml;q=0.9,"
                    "image/avif,"
                    "image/webp,"
                    "*/*;q=0.8"
                ),
                "Accept-Language": (
                    "id-ID,id;q=0.9,"
                    "en-US;q=0.8,en;q=0.7"
                ),
                "Connection": "keep-alive",
            }
        )

        self.current_url = ""

        self.token: str | None = None

        self.trace: list[str] = []

        self.logged_in = False

        self.last_login_at = 0.0

        # Pengolahan-only cache of parameter-specific BBWS set_sensordash
        # tokens. Kept on this authenticated client so a forced re-login can
        # invalidate all entries safely.
        self._sensor_token_cache: dict[tuple[str, str], tuple[float, str, str]] = {}

    # --------------------------------------------------------
    # TRACE
    # --------------------------------------------------------

    def note(
        self,
        message: str,
    ) -> None:

        self.trace.append(message)

        if len(self.trace) > 100:
            self.trace = self.trace[-100:]

    # --------------------------------------------------------
    # GET
    # --------------------------------------------------------

    def _get(
        self,
        url: str,
        *,
        referer: str | None = None,
    ) -> requests.Response:

        headers = {}

        if referer:
            headers["Referer"] = referer

        response = self.session.get(
            url,
            timeout=TIMEOUT,
            allow_redirects=True,
            headers=headers,
        )

        response.raise_for_status()

        self.current_url = response.url

        return response

    # --------------------------------------------------------
    # POST
    # --------------------------------------------------------

    def _post(
        self,
        url: str,
        data: dict[str, str],
        *,
        referer: str | None = None,
    ) -> requests.Response:

        headers = {
            "Content-Type": (
                "application/x-www-form-urlencoded"
            ),
            "Origin": BASE_URL,
        }

        if referer:
            headers["Referer"] = referer

        response = self.session.post(
            url,
            data=data,
            timeout=TIMEOUT,
            allow_redirects=True,
            headers=headers,
        )

        response.raise_for_status()

        self.current_url = response.url

        return response

    # --------------------------------------------------------
    # LOGIN
    # --------------------------------------------------------

    def login(
        self,
        force: bool = False,
    ) -> None:

        if (
            self.logged_in
            and self.token
            and not force
        ):
            return

        if not USERNAME or not PASSWORD:

            raise RuntimeError(
                "BEACON_USERNAME dan BEACON_PASSWORD "
                "belum dikonfigurasi."
            )

        if USERNAME.startswith("ISI_"):

            raise RuntimeError(
                "Isi BEACON_USERNAME dan BEACON_PASSWORD "
                "di Environment Variable/config.py."
            )

        if force:

            self.logged_in = False
            self.token = None
            self.current_url = ""
            self._sensor_token_cache.clear()

            try:
                self.session.close()
            except Exception:
                pass

            self.session = requests.Session()

            self.session.headers.update(
                {
                    "User-Agent": (
                        "Mozilla/5.0 "
                        "(Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 "
                        "Chrome/151.0 Safari/537.36"
                    ),
                    "Accept": (
                        "text/html,"
                        "application/xhtml+xml,"
                        "application/xml;q=0.9,"
                        "*/*;q=0.8"
                    ),
                    "Accept-Language": (
                        "id-ID,id;q=0.9,"
                        "en-US;q=0.8"
                    ),
                    "Connection": "keep-alive",
                }
            )

        self.note("GET /login")

        page = self._get(
            LOGIN_URL
        )

        action, payload, username_name, password_name = (
            find_login_form(page.text)
        )

        payload[username_name] = USERNAME
        payload[password_name] = PASSWORD

        self.note(
            "POST login: "
            f"username={username_name}, "
            f"password={password_name}"
        )

        response = self._post(
            action,
            payload,
            referer=LOGIN_URL,
        )

        if looks_like_login(
            response.text,
            response.url,
        ):

            raise RuntimeError(
                "Login BBWS gagal atau server "
                "mengembalikan halaman login lagi."
            )

        self.logged_in = True
        self.token = None
        self.last_login_at = time.time()

        self.note(
            "Login berhasil: "
            f"{response.status_code} → "
            f"{response.url}"
        )

    # --------------------------------------------------------
    # ANALYSIS PAGE
    # --------------------------------------------------------

    def analysis_page(self) -> str:

        if not self.logged_in:

            self.login()

        if self.token:

            return ""

        response = self._get(
            ANALISA_URL,
            referer=self.current_url or LOGIN_URL,
        )

        if looks_like_login(
            response.text,
            response.url,
        ):

            self.logged_in = False
            self.token = None

            raise RuntimeError(
                "Sesi login tidak aktif "
                "saat membuka /analisa."
            )

        self.token = extract_token(
            response.text,
            response.url,
        )

        self.note(
            f"GET /analisa: "
            f"{response.status_code} → "
            f"{response.url}"
        )

        if not self.token:

            raise RuntimeError(
                "Token analisa tidak ditemukan "
                "pada halaman /analisa."
            )

        self.note(
            "Token analisa ditemukan."
        )

        return response.text

    # --------------------------------------------------------
    # SET STATE
    # --------------------------------------------------------

    def set_state(
        self,
        **fields: str,
    ) -> str:

        if not self.token:

            self.analysis_page()

        payload = {
            "token": self.token
        }

        payload.update(
            {
                k: v
                for k, v in fields.items()
                if v not in {
                    None,
                    "",
                }
            }
        )

        self.note(
            "POST /analisa/set_token: "
            + ", ".join(
                f"{k}={v}"
                for k, v in fields.items()
            )
        )

        response = self._post(
            SET_TOKEN_URL,
            payload,
            referer=self.current_url or ANALISA_URL,
        )

        if looks_like_login(
            response.text,
            response.url,
        ):

            self.logged_in = False
            self.token = None

            raise RuntimeError(
                "Sesi BBWS kedaluwarsa."
            )

        self.current_url = response.url

        new_token = extract_token(
            response.text,
            response.url,
        )

        if new_token:

            self.token = new_token

        return response.text

    # --------------------------------------------------------
    # PARALLEL DATA CHUNK HELPERS
    # --------------------------------------------------------

    def _clone_http_pool(self) -> requests.Session:
        """Clone authenticated cookies/headers into an isolated pool.

        requests.Session is not shared between worker threads. The Beacon token
        itself is stable after set_token, so independent /data_chunk requests can
        safely run in parallel using cloned cookie jars.
        """
        client = requests.Session()
        client.headers.update(dict(self.session.headers))
        client.cookies.update(self.session.cookies)
        return client

    def _fetch_data_chunk(
        self,
        token: str,
        start: datetime,
        end: datetime,
    ) -> list[list[str]]:
        client = self._clone_http_pool()
        files = {
            "token": (None, token),
            "start": (None, start.strftime("%Y-%m-%d %H:%M:%S")),
            "end": (None, end.strftime("%Y-%m-%d %H:%M:%S")),
        }
        response = client.post(
            DATA_CHUNK_URL,
            files=files,
            headers={
                "Accept": "application/json, text/plain, */*",
                "Origin": BASE_URL,
                "Referer": self.current_url or f"{BASE_URL}/analisa/data/{token}",
            },
            timeout=max(TIMEOUT, 60),
            allow_redirects=True,
        )
        response.raise_for_status()
        if looks_like_login(response.text, response.url):
            raise RuntimeError("Sesi BBWS kedaluwarsa saat mengambil data_chunk.")
        try:
            payload = response.json()
        except ValueError as exc:
            raise RuntimeError("Response data_chunk Beacon bukan JSON.") from exc

        if not isinstance(payload, dict):
            raise RuntimeError("Payload data_chunk Beacon tidak valid.")

        status = clean_text(str(payload.get("status", ""))).lower()
        raw_data = payload.get("data")
        table = payload.get("data_tabel")
        if not isinstance(raw_data, list):
            raw_data = []
        if not isinstance(table, list):
            table = []

        # Beacon can legitimately return an empty series outside a logger's
        # recording lifetime. Treat that as EMPTY, not a failed request.
        if status and status not in {"ok", "success"} and not raw_data and not table:
            msg = clean_text(str(payload.get("message") or payload.get("error") or status))
            if any(k in msg.lower() for k in ("no data", "tidak ada data", "empty", "kosong")):
                return []
            raise RuntimeError(f"data_chunk Beacon gagal: {msg}")

        # IMPORTANT: /analisa/data_chunk exposes the canonical time series in
        # ``data`` as [epoch_milliseconds, value]. ``data_tabel`` is a display
        # helper used by Beacon's page and, for some responses, its ``waktu``
        # value is only a row/index-like number. Passing that number to the
        # frontend date parser produced the 1899-12-30 / 1900-01-xx dates seen
        # in Pengolahan V26. Monitoring already uses ``data`` correctly, so the
        # processing adapter now follows the same source of truth.
        rows: list[list[str]] = []
        for item in raw_data:
            if not isinstance(item, (list, tuple)) or len(item) < 2:
                continue
            try:
                # Use utcfromtimestamp intentionally: Beacon's chart/table
                # clock is represented directly by this epoch value. This also
                # avoids host/Vercel timezone differences.
                dt = datetime.utcfromtimestamp(float(item[0]) / 1000.0)
            except (TypeError, ValueError, OSError, OverflowError):
                continue
            value = item[1]
            if value is None or clean_text(str(value)) == "":
                continue
            rows.append([
                dt.strftime("%Y-%m-%d %H:%M:%S"),
                clean_text(str(value)),
            ])

        if rows:
            return rows

        # Compatibility fallback only for an upstream response that does not
        # contain ``data``. Accept data_tabel timestamps only when they parse as
        # a real datetime, preventing numeric row indexes from becoming Excel
        # epoch dates in the frontend.
        for item in table:
            if not isinstance(item, dict):
                continue
            stamp = clean_text(str(item.get("waktu") or item.get("time") or ""))
            value = item.get("dta")
            if value is None:
                value = item.get("value")
            parsed_stamp = _parse_any_datetime(stamp)
            if parsed_stamp is None or value is None or clean_text(str(value)) == "":
                continue
            rows.append([
                parsed_stamp.strftime("%Y-%m-%d %H:%M:%S"),
                clean_text(str(value)),
            ])
        return rows

    # --------------------------------------------------------
    # DISCOVER PARAMETERS
    # --------------------------------------------------------

    def discover_parameters(
        self,
        id_logger: str,
    ) -> list[dict[str, str]]:

        # Jika parameter sudah tersedia di cache,
        # JANGAN mengakses BBWS lagi.

        cached = PARAMETER_CACHE.get(
            str(id_logger)
        )

        if cached:

            age = time.time() - cached[0]

            if age < PARAMETER_CACHE_TTL:

                return cached[1]

        # ----------------------------------------------------
        # Live discovery
        # ----------------------------------------------------

        self.analysis_page()

        html = self.set_state(
            id_logger=id_logger
        )

        params = extract_options(
            html,
            "select-parameter",
        )

        if not params and self.current_url:

            response = self._get(
                self.current_url,
                referer=ANALISA_URL,
            )

            html = response.text

            params = extract_options(
                html,
                "select-parameter",
            )

        if not params:

            raise RuntimeError(
                "Tidak ada parameter yang "
                f"ditampilkan server untuk logger "
                f"{id_logger}."
            )

        result = parameter_with_metadata(
            params
        )

        PARAMETER_CACHE[
            str(id_logger)
        ] = (
            time.time(),
            result,
        )

        _save_parameter_catalog()

        return result

    # --------------------------------------------------------
    # PREPARE HISTORICAL
    # --------------------------------------------------------

    def _prepare_once(
        self,
        id_logger: str,
        id_param: str,
        dari: str,
        sampai: str,
        mode: str = "range",
    ) -> str:

        if not self.logged_in:

            self.login()

        if not self.token:

            self.analysis_page()

        payload = {
            "id_logger": id_logger,
            "id_param": id_param,
            "mode": mode,
            "dari": dari,
            "sampai": sampai,
        }

        self.note(
            "POST combined set_token: "
            f"logger={id_logger}, "
            f"param={id_param}, "
            f"mode={mode}, "
            f"dari={dari}, "
            f"sampai={sampai}"
        )

        # ====================================================
        # FAST PATH
        # ====================================================

        try:

            final_html = self.set_state(
                **payload
            )

            if (
                "tbl_exporttable_to_xls"
                in final_html
            ):

                return final_html

            # Server dapat mengembalikan halaman
            # tanpa tabel tetapi bukan error HTTP.
            #
            # Coba GET halaman hasil satu kali.

            candidates = []

            if self.current_url:
                candidates.append(
                    self.current_url
                )

            if self.token:
                candidates.append(
                    f"{BASE_URL}/analisa/data/{self.token}"
                )

            for candidate in dict.fromkeys(
                candidates
            ):

                response = self._get(
                    candidate,
                    referer=(
                        self.current_url
                        or ANALISA_URL
                    ),
                )

                if looks_like_login(
                    response.text,
                    response.url,
                ):

                    self.logged_in = False
                    self.token = None

                    raise RuntimeError(
                        "Sesi BBWS kedaluwarsa "
                        "saat mengambil data historis."
                    )

                if (
                    "tbl_exporttable_to_xls"
                    in response.text
                ):

                    self.current_url = (
                        response.url
                    )

                    return response.text

            # ------------------------------------------------
            # Tidak ada tabel.
            #
            # Ini BELUM tentu data kosong.
            # Jangan menyatakan "tidak ada data".
            # ------------------------------------------------

            raise RuntimeError(
                "Server BBWS tidak mengembalikan "
                "tabel historis."
            )

        except requests.HTTPError as exc:

            status = (
                exc.response.status_code
                if exc.response is not None
                else None
            )

            # 520/522/5xx = SERVER ERROR.
            # Jangan dianggap data kosong.

            if status is not None:

                raise RuntimeError(
                    f"Server BBWS mengembalikan "
                    f"HTTP {status}."
                ) from exc

            raise

        except requests.RequestException:

            raise

    # --------------------------------------------------------
    # FETCH HISTORICAL
    # --------------------------------------------------------

    @staticmethod
    def _logger_supports_data_chunk(id_logger: str) -> bool:
        """Return True only for Beacon assets owned by BBWS.

        The upstream ``/analisa/data_chunk`` endpoint is intentionally limited
        to logger IDs carrying the ``_bbws`` asset marker. PSDA/non-BBWS assets
        use the normal HTML analysis result instead. This distinction is also
        visible in the Beacon UI: examples include ``10090_psda`` (non-BBWS)
        versus ``10044_bbws`` (BBWS).
        """
        return "_bbws" in clean_text(id_logger).lower()

    def _fetch_historical_html_chunks(
        self,
        id_logger: str,
        id_param: str,
        start: datetime,
        end: datetime,
        *,
        mode: str,
    ) -> tuple[list[str], list[list[str]], str]:
        """Legacy Beacon HTML path for PSDA/non-BBWS assets.

        This is the same mechanism used by the previous working application:
        each range is applied through ``/analisa/set_token`` and the resulting
        ``tbl_exporttable_to_xls`` table is parsed. Requests remain capped by
        ``BEACON_CHUNK_DAYS``; a range-like failure is retried with a smaller
        sub-range. Empty edge ranges are skipped rather than aborting the whole
        requested period.
        """
        fmt = "%Y-%m-%d %H:%M"
        all_headers: list[str] = []
        all_rows: list[list[str]] = []
        title = "Data Telemetri BBWS Serayu Opak"
        seen_keys: set[str] = set()
        cursor = start

        primary_days = max(1, min(25, int(BEACON_CHUNK_DAYS)))
        retry_days = max(1, min(12, primary_days))
        chunk_sizes = [primary_days]
        if retry_days < primary_days:
            chunk_sizes.append(retry_days)

        while cursor <= end:
            successful = False
            last_error: Exception | None = None

            for attempt_index, chunk_days in enumerate(chunk_sizes, start=1):
                candidate_end = min(
                    end,
                    cursor + timedelta(days=chunk_days) - timedelta(minutes=1),
                )
                self.note(
                    "Beacon HTML non-BBWS: "
                    f"attempt {attempt_index}/{len(chunk_sizes)} "
                    f"({chunk_days} hari)"
                )

                try:
                    chunk_html = self._prepare_once(
                        id_logger=id_logger,
                        id_param=id_param,
                        dari=cursor.strftime(fmt),
                        sampai=candidate_end.strftime(fmt),
                        mode=mode,
                    )
                    headers, rows, chunk_title = extract_table(chunk_html)

                    if not all_headers:
                        all_headers = headers
                    if chunk_title:
                        title = chunk_title

                    for row in rows:
                        if not row:
                            continue
                        # Timestamp is the stable dedupe key for a selected
                        # parameter. Keep the old application's behaviour.
                        key = clean_text(str(row[0])) if row else ""
                        if not key:
                            key = "|".join(str(value) for value in row)
                        if key and key in seen_keys:
                            continue
                        if key:
                            seen_keys.add(key)
                        all_rows.append(row)

                    cursor = candidate_end + timedelta(minutes=1)
                    successful = True
                    break

                except EmptyHistoricalData:
                    # Sparse beginning/end periods are valid. Advance the
                    # cursor so a later chunk can still contain telemetry.
                    cursor = candidate_end + timedelta(minutes=1)
                    successful = True
                    break

                except RuntimeError as exc:
                    last_error = exc
                    message = str(exc).lower()
                    range_keywords = (
                        "range", "rentang", "terlalu panjang", "maximum",
                        "maksimum", "period too long", "too long", "too many",
                        "limit periode", "limit period", "timeout", "timed out",
                    )
                    can_retry_smaller = (
                        attempt_index < len(chunk_sizes)
                        and any(keyword in message for keyword in range_keywords)
                    )
                    if can_retry_smaller:
                        continue
                    raise

                except requests.RequestException as exc:
                    last_error = exc
                    if attempt_index < len(chunk_sizes):
                        continue
                    raise RuntimeError(
                        "Gagal terhubung ke server Beacon non-BBWS: "
                        f"{exc}"
                    ) from exc

            if not successful:
                if last_error:
                    raise RuntimeError(
                        "Gagal mengambil data Beacon non-BBWS: "
                        f"{last_error}"
                    ) from last_error
                raise RuntimeError("Gagal mengambil data Beacon non-BBWS.")

        if not all_rows:
            raise EmptyHistoricalData("Tidak ada data pada rentang yang diminta.")

        all_rows.sort(key=lambda row: row[0] if row else "")
        return all_headers, all_rows, title

    def _bbws_fast_selector_url(self, id_logger: str, id_param: str) -> str:
        lid = clean_text(str(id_logger))
        pid = clean_text(str(id_param))
        lower = lid.lower()
        if "_bbws" not in lower:
            raise RuntimeError("Fast selector hanya didukung untuk aset BBWS.")
        grp_match = re.search(r"_bbws_(\d+)$", lower)
        suffix = f"&grp={quote_plus(grp_match.group(1))}" if grp_match else ""
        return (
            f"{BASE_URL}/analisa/set_sensordash?"
            f"id_param={quote_plus(pid + '_bbws')}{suffix}"
        )

    def _prepare_bbws_fast_token(
        self,
        id_logger: str,
        id_param: str,
        *,
        force_refresh: bool = False,
    ) -> tuple[str, str]:
        """Get a parameter-specific BBWS token without rendering analysis HTML.

        Monitoring proved that /analisa/set_sensordash redirects directly to
        /analisa/data/<token>. Pengolahan only needs that token for data_chunk,
        so cold requests can skip GET /analisa + POST set_token + a full result
        page render. The legacy set_state path remains the fallback.
        """
        if not self.logged_in:
            self.login()

        key = (clean_text(str(id_logger)), clean_text(str(id_param)))
        now = time.time()
        if not force_refresh:
            cached = self._sensor_token_cache.get(key)
            if cached and now - cached[0] < BEACON_PROCESS_TOKEN_TTL:
                _created_at, token, current_url = cached
                self.token = token
                self.current_url = current_url or f"{BASE_URL}/analisa/data/{token}"
                return token, "hit"
        else:
            self._sensor_token_cache.pop(key, None)

        selector_url = self._bbws_fast_selector_url(id_logger, id_param)

        def select_once() -> tuple[str, str | None]:
            referer = self.current_url or f"{BASE_URL}/beranda"
            response = self.session.get(
                selector_url,
                headers={"Referer": referer},
                timeout=max(TIMEOUT, 60),
                allow_redirects=False,
            )
            response.raise_for_status()
            location = clean_text(str(response.headers.get("Location") or ""))
            if location:
                absolute = urljoin(BASE_URL + "/", location)
                if "/login" in urlparse(absolute).path.lower():
                    raise RuntimeError("Sesi BBWS kedaluwarsa saat set_sensordash.")
                match = re.search(r"/analisa/data/([^/?#]+)", absolute)
                if match:
                    return absolute, match.group(1)

            # Defensive fallback only when the vendor changes redirect format.
            followed = self.session.get(
                selector_url,
                headers={"Referer": referer},
                timeout=max(TIMEOUT, 60),
                allow_redirects=True,
            )
            followed.raise_for_status()
            if looks_like_login(followed.text, followed.url):
                raise RuntimeError("Sesi BBWS kedaluwarsa saat set_sensordash.")
            return followed.url, extract_token(followed.text, followed.url)

        try:
            selected_url, token = select_once()
        except Exception as first_exc:
            # One auth refresh protects warm Vercel/local instances whose cookie
            # expired while the in-memory client itself was still alive.
            try:
                self.login(force=True)
                selected_url, token = select_once()
            except Exception:
                raise first_exc

        if not token:
            raise RuntimeError("Token Beacon tidak ditemukan setelah set_sensordash.")

        current_url = selected_url or f"{BASE_URL}/analisa/data/{token}"
        self.token = token
        self.current_url = current_url
        self._sensor_token_cache[key] = (time.time(), token, current_url)
        return token, "miss"

    def _prepare_bbws_legacy_token(
        self,
        id_logger: str,
        id_param: str,
        start: datetime,
        end: datetime,
        mode: str,
    ) -> str:
        """Original set_token preparation retained as a compatibility fallback."""
        fmt = "%Y-%m-%d %H:%M"
        if not self.logged_in:
            self.login()
        if not self.token:
            self.analysis_page()
        self.set_state(
            id_logger=id_logger,
            id_param=id_param,
            mode=mode,
            dari=start.strftime(fmt),
            sampai=end.strftime(fmt),
        )
        if not self.token:
            raise RuntimeError("Token data Beacon tidak ditemukan setelah set_token.")
        return self.token

    def _fetch_historical_data_chunks(
        self,
        id_logger: str,
        id_param: str,
        start: datetime,
        end: datetime,
        *,
        mode: str,
        parameter_name: str,
        parallel_workers: int | None,
    ) -> tuple[list[str], list[list[str]], str]:
        """Optimized ``/data_chunk`` path for ``*_bbws`` Beacon assets."""
        fmt = "%Y-%m-%d %H:%M"

        # Fast path: get/reuse the parameter token directly from set_sensordash.
        # If Beacon changes that route, fall back to the proven set_token flow.
        token_status = "legacy"
        try:
            token, token_status = self._prepare_bbws_fast_token(id_logger, id_param)
        except Exception:
            token = self._prepare_bbws_legacy_token(
                id_logger=id_logger, id_param=id_param, start=start, end=end, mode=mode
            )

        chunk_days = max(1, min(25, int(BEACON_CHUNK_DAYS)))
        chunks: list[tuple[datetime, datetime]] = []
        cursor = start
        while cursor <= end:
            candidate_end = min(
                end,
                cursor + timedelta(days=chunk_days) - timedelta(minutes=1),
            )
            chunks.append((cursor, candidate_end))
            cursor = candidate_end + timedelta(minutes=1)

        workers = parallel_workers if parallel_workers is not None else BEACON_PARALLEL_WORKERS
        workers = max(1, min(int(workers), len(chunks) if chunks else 1, 4))
        results: dict[int, list[list[str]]] = {}
        failures: dict[int, Exception] = {}

        def run_chunk(index: int, period: tuple[datetime, datetime]) -> tuple[int, list[list[str]]]:
            last_exc: Exception | None = None
            for attempt in range(2):
                try:
                    return index, self._fetch_data_chunk(token, period[0], period[1])
                except Exception as exc:
                    last_exc = exc
                    if attempt == 0:
                        time.sleep(0.15)
            assert last_exc is not None
            raise last_exc

        if workers > 1 and len(chunks) > 1:
            with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="beacon") as executor:
                future_map = {executor.submit(run_chunk, i, ch): i for i, ch in enumerate(chunks)}
                for future in as_completed(future_map):
                    idx = future_map[future]
                    try:
                        _, rows = future.result()
                        results[idx] = rows
                    except Exception as exc:
                        failures[idx] = exc
        else:
            for i, ch in enumerate(chunks):
                try:
                    _, rows = run_chunk(i, ch)
                    results[i] = rows
                except Exception as exc:
                    failures[i] = exc

        # A warm cached token can expire independently of the login cookie.
        # Refresh it once before applying the smaller-range fallback.
        if failures and token_status == "hit":
            try:
                token, _ = self._prepare_bbws_fast_token(
                    id_logger, id_param, force_refresh=True
                )
                retry_indexes = sorted(list(failures))
                for idx in retry_indexes:
                    try:
                        _, rows = run_chunk(idx, chunks[idx])
                        results[idx] = rows
                        failures.pop(idx, None)
                    except Exception as exc:
                        failures[idx] = exc
            except Exception:
                pass

        # Timeout/range fallback: retry a failed BBWS chunk as <=12-day pieces.
        for idx in sorted(list(failures)):
            ch_start, ch_end = chunks[idx]
            subrows: list[list[str]] = []
            subcur = ch_start
            try:
                while subcur <= ch_end:
                    subend = min(ch_end, subcur + timedelta(days=min(12, chunk_days)) - timedelta(minutes=1))
                    subrows.extend(self._fetch_data_chunk(token, subcur, subend))
                    subcur = subend + timedelta(minutes=1)
                results[idx] = subrows
                failures.pop(idx, None)
            except Exception as exc:
                failures[idx] = exc

        if failures:
            detail = "; ".join(f"bagian {i+1}: {e}" for i, e in sorted(failures.items()))
            raise RuntimeError(f"Sebagian data Beacon gagal setelah retry: {detail}")

        all_rows: list[list[str]] = []
        seen: set[tuple[str, str]] = set()
        for idx in range(len(chunks)):
            for row in results.get(idx, []):
                if not row:
                    continue
                key = (str(row[0]), str(row[1] if len(row) > 1 else ""))
                if key in seen:
                    continue
                seen.add(key)
                all_rows.append(row)

        if not all_rows:
            raise EmptyHistoricalData("Tidak ada data pada rentang yang diminta.")
        all_rows.sort(key=lambda row: row[0] if row else "")
        return ["Waktu", parameter_name or "Data"], all_rows, "Data Telemetri BBWS Serayu Opak"

    def fetch_historical(
        self,
        id_logger: str,
        id_param: str,
        dari: str,
        sampai: str,
        mode: str = "range",
        period_mode: str = "",
        parameter_name: str = "Data",
        parallel_workers: int | None = None,
    ) -> tuple[list[str], list[list[str]], str]:
        """Fetch Beacon history with the correct upstream path per asset owner.

        ``*_bbws`` loggers use the fast JSON ``/data_chunk`` endpoint. Beacon
        ``*_psda`` / other non-BBWS assets keep the proven HTML-table path,
        because the upstream chunk endpoint rejects those assets with
        "Chunk hanya didukung untuk aset BBWS".
        """
        fmt = "%Y-%m-%d %H:%M"
        try:
            start = datetime.strptime(dari.strip(), fmt)
            end = datetime.strptime(sampai.strip(), fmt)
        except ValueError as exc:
            raise RuntimeError("Format tanggal sumber harus YYYY-MM-DD HH:MM.") from exc
        if end < start:
            raise RuntimeError("Tanggal mulai tidak boleh melebihi tanggal akhir.")

        now = datetime.now().replace(second=0, microsecond=0)
        if clean_text(period_mode).lower() == "year" and start.year == now.year and end > now:
            end = now
            if end < start:
                raise EmptyHistoricalData("Tidak ada data pada rentang yang diminta.")

        if self._logger_supports_data_chunk(id_logger):
            return self._fetch_historical_data_chunks(
                id_logger=id_logger,
                id_param=id_param,
                start=start,
                end=end,
                mode=mode,
                parameter_name=parameter_name,
                parallel_workers=parallel_workers,
            )

        return self._fetch_historical_html_chunks(
            id_logger=id_logger,
            id_param=id_param,
            start=start,
            end=end,
            mode=mode,
        )

    # --------------------------------------------------------
    # PREPARE
    # --------------------------------------------------------

    def prepare(
        self,
        id_logger: str,
        id_param: str,
        dari: str,
        sampai: str,
        mode: str = "range",
    ) -> str:

        return self._prepare_once(
            id_logger,
            id_param,
            dari,
            sampai,
            mode,
        )


# ============================================================
# POSITION HELPERS
# ============================================================

def source_positions() -> list[
    dict[str, str]
]:

    out = []

    seen = set()

    for p in FALLBACK_POSITIONS:

        logger_id = clean_text(
            p.get(
                "id_logger",
                "",
            )
        )

        if (
            not logger_id
            or logger_id in seen
        ):
            continue

        seen.add(logger_id)

        full_name = clean_text(
            p.get(
                "name",
                logger_id,
            )
        )

        out.append(
            {
                "id_logger": logger_id,
                "name": station_alias(
                    "beacon",
                    logger_id,
                    display_position_name(full_name),
                ),
                "full_name": full_name,
            }
        )

    return sorted(
        out,
        key=lambda x: (
            x["name"].lower(),
            x["id_logger"],
        ),
    )


def _ambiguous_positions(
    data_type: str,
) -> list[dict[str, str]]:

    out = []

    for pos in source_positions():

        raw = pos[
            "full_name"
        ].lower()

        is_awlr = (
            "awlr" in raw
        )

        is_awr = (
            re.search(
                r"\bawr\b",
                raw,
            )
            is not None
            and not is_awlr
        )

        is_afmr = (
            "afmr" in raw
        )

        if data_type in {
            "rain",
            "tma",
        } and (
            is_awlr
            or is_awr
            or is_afmr
        ):

            out.append(pos)

    return out


def _quick_positions_for_data_type(
    data_type: str,
) -> list[dict[str, str]]:

    catalog_ids = set(
        PARAMETER_CACHE.keys()
    )

    result = []

    target = (
        "rainfall"
        if data_type == "rain"
        else "water_level"
    )

    for pos in source_positions():

        raw = pos[
            "full_name"
        ].lower()

        is_awlr = (
            "awlr" in raw
        )

        is_awr = (
            re.search(
                r"\bawr\b",
                raw,
            )
            is not None
            and not is_awlr
        )

        is_arr = (
            "arr" in raw
        )

        is_afmr = (
            "afmr" in raw
        )

        is_aws = (
            re.search(
                r"\baws\b",
                raw,
            )
            is not None
        )

        is_climate = (
            "klimatologi"
            in raw
        )

        logger_id = pos[
            "id_logger"
        ]

        # ----------------------------------------------------
        # Jika catalog tersedia:
        # GUNAKAN catalog.
        # ----------------------------------------------------

        if logger_id in catalog_ids:

            params = PARAMETER_CACHE[
                logger_id
            ][1]

            if any(
                p["type"] == target
                for p in params
            ):

                result.append(pos)

            continue

        # ----------------------------------------------------
        # Belum ada catalog:
        # heuristic.
        # ----------------------------------------------------

        if data_type == "rain":

            if (
                is_arr
                or is_aws
                or is_afmr
                or is_climate
            ):

                result.append(pos)

        else:

            if (
                is_awlr
                or is_awr
                or is_afmr
            ):

                result.append(pos)

    return sorted(
        result,
        key=lambda x: (
            x["name"].lower(),
            x["id_logger"],
        ),
    )


# ============================================================
# BACKGROUND CATALOG WARMING
# ============================================================

def _warm_position_catalog(
    data_type: str,
) -> None:

    with CATALOG_LOCK:

        if (
            data_type
            in CATALOG_WARMING
        ):

            return

        CATALOG_WARMING.add(
            data_type
        )

    def worker():

        try:

            candidates = (
                _ambiguous_positions(
                    data_type
                )
            )

            client = get_bbws_client()

            for pos in candidates:

                logger_id = pos[
                    "id_logger"
                ]

                if (
                    logger_id
                    in PARAMETER_CACHE
                ):

                    continue

                try:

                    with BBWS_CLIENT_LOCK:

                        params = (
                            client.discover_parameters(
                                logger_id
                            )
                        )

                    PARAMETER_CACHE[
                        logger_id
                    ] = (
                        time.time(),
                        params,
                    )

                    _save_parameter_catalog()

                except Exception as exc:

                    print(
                        "Catalog warming gagal "
                        f"{logger_id}: {exc}"
                    )

                    continue

        finally:

            with CATALOG_LOCK:

                CATALOG_WARMING.discard(
                    data_type
                )

    threading.Thread(
        target=worker,
        daemon=True,
    ).start()


def _catalog_ready(
    data_type: str,
) -> bool:

    candidates = (
        _ambiguous_positions(
            data_type
        )
    )

    return all(
        p["id_logger"]
        in PARAMETER_CACHE
        for p in candidates
    )


def positions_for_data_type(
    data_type: str,
) -> list[dict[str, str]]:

    data_type = (
        clean_text(data_type)
        or "rain"
    )

    now = time.time()

    cached = POSITION_TYPE_CACHE.get(
        data_type
    )

    if cached:

        age = now - cached[0]

        if age < PARAMETER_CACHE_TTL:

            return cached[1]

    result = _quick_positions_for_data_type(
        data_type
    )

    # Jangan menunggu warming.
    #
    # UI langsung mendapatkan hasil
    # dari catalog/heuristic.

    _warm_position_catalog(
        data_type
    )

    if _catalog_ready(
        data_type
    ):

        POSITION_TYPE_CACHE[
            data_type
        ] = (
            now,
            result,
        )

    return result


# ============================================================
# TABLE PARSER
# ============================================================

def extract_table(
    html: str,
) -> tuple[
    list[str],
    list[list[str]],
    str,
]:

    soup = parse_html(html)

    table = soup.find(
        "table",
        id="tbl_exporttable_to_xls",
    )

    if table is None:

        # Tidak ada tabel.
        #
        # Kita tidak boleh langsung menganggap
        # server error atau data kosong.
        #
        # Cari indikasi empty result.

        text = clean_text(
            soup.get_text(
                " ",
                strip=True,
            )
        ).lower()

        empty_keywords = [
            "tidak ada data",
            "data tidak ditemukan",
            "no data",
            "data kosong",
            "record not found",
            "tidak ditemukan",
        ]

        if any(
            keyword in text
            for keyword in empty_keywords
        ):

            raise EmptyHistoricalData(
                "Tidak ada data pada "
                "rentang yang diminta."
            )

        raise RuntimeError(
            "Tabel data historis tidak ditemukan."
        )

    headers: list[str] = []

    for tr in table.find_all("tr"):

        vals = [
            clean_text(
                x.get_text(
                    " ",
                    strip=True,
                )
            )
            for x in tr.find_all(
                [
                    "th",
                    "td",
                ]
            )
        ]

        if (
            any(
                v.lower() == "waktu"
                for v in vals
            )
            or any(
                "tinggi muka air"
                in v.lower()
                for v in vals
            )
        ):

            headers = vals

    if not headers:

        thead = table.find(
            "thead"
        )

        if thead:

            rows = thead.find_all(
                "tr"
            )

            if rows:

                headers = [
                    clean_text(
                        x.get_text(
                            " ",
                            strip=True,
                        )
                    )
                    for x in rows[-1].find_all(
                        [
                            "th",
                            "td",
                        ]
                    )
                ]

    rows: list[
        list[str]
    ] = []

    body = table.find(
        "tbody"
    )

    if body:

        for tr in body.find_all(
            "tr"
        ):

            vals = [
                clean_text(
                    x.get_text(
                        " ",
                        strip=True,
                    )
                )
                for x in tr.find_all(
                    [
                        "td",
                        "th",
                    ]
                )
            ]

            if vals:
                rows.append(vals)

    if not rows:

        raise EmptyHistoricalData(
            "Tidak ada data pada "
            "rentang yang diminta."
        )

    if not headers:

        headers = [
            f"Kolom {i + 1}"
            for i in range(
                len(rows[0])
            )
        ]

    h5 = table.find(
        "h5"
    )

    title = (
        clean_text(
            h5.get_text(
                " ",
                strip=True,
            )
        )
        if h5
        else "Data Telemetri BBWS Serayu Opak"
    )

    return (
        headers,
        rows,
        title,
    )


# ============================================================
# EXCEL VALUE
# ============================================================

def value_for_excel(
    value: str,
) -> Any:

    value = clean_text(
        value
    )

    if not value:
        return None

    if re.match(
        r"^\d{4}-\d{2}-\d{2}$",
        value,
    ):

        try:

            return datetime.strptime(
                value,
                "%Y-%m-%d",
            ).date()

        except ValueError:

            pass

    if re.match(
        r"^\d{4}-\d{2}-\d{2} "
        r"\d{2}:\d{2}:\d{2}$",
        value,
    ):

        try:

            return datetime.strptime(
                value,
                "%Y-%m-%d %H:%M:%S",
            )

        except ValueError:

            pass

    m = re.match(
        r"^([-+]?\d+(?:[\.,]\d+)?)"
        r"\s*(?:[A-Za-zµ³²/\-]+)?$",
        value,
    )

    if m:

        try:

            return float(
                m.group(1).replace(
                    ",",
                    ".",
                )
            )

        except ValueError:

            pass

    return value


# ============================================================
# XLSX
# ============================================================

def build_xlsx(
    pos_name: str,
    parameter_name: str,
    dari: str,
    sampai: str,
    title: str,
    headers: list[str],
    rows: list[list[str]],
) -> io.BytesIO:

    wb = Workbook()

    ws = wb.active

    ws.title = "DATA"

    ws.append(headers)

    for row in rows:

        padded = (
            row
            + [""] * max(
                0,
                len(headers) - len(row),
            )
        )

        ws.append(
            [
                value_for_excel(v)
                for v in padded[
                    :len(headers)
                ]
            ]
        )

    fill = PatternFill(
        "solid",
        fgColor="303481",
    )

    for cell in ws[1]:

        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    for i, column in enumerate(
        ws.columns,
        start=1,
    ):

        max_len = max(
            len(
                str(cell.value)
            )
            if cell.value is not None
            else 0
            for cell in column
        )

        ws.column_dimensions[
            get_column_letter(i)
        ].width = min(
            max(max_len + 2, 12),
            32,
        )

    for row in ws.iter_rows(
        min_row=2
    ):

        if row and isinstance(row[0].value, datetime):

            row[0].number_format = "yyyy-mm-dd hh:mm:ss"

        elif row and isinstance(row[0].value, date):

            row[0].number_format = "yyyy-mm-dd"

    info = wb.create_sheet(
        "INFO"
    )

    for item in [
        (
            "Sumber",
            BASE_URL,
        ),
        (
            "Pos",
            pos_name,
        ),
        (
            "Parameter",
            parameter_name,
        ),
        (
            "Dari",
            dari,
        ),
        (
            "Sampai",
            sampai,
        ),
        (
            "Judul sumber",
            title,
        ),
        (
            "Dibuat",
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        ),
        (
            "Metode",
            (
                "Login → /analisa → "
                "/analisa/set_token → "
                "HTML historis → "
                "parsing tabel → XLSX"
            ),
        ),
    ]:

        info.append(item)

    info.column_dimensions[
        "A"
    ].width = 24

    info.column_dimensions[
        "B"
    ].width = 90

    for cell in info[1]:

        cell.font = Font(
            bold=True
        )

    out = io.BytesIO()

    wb.save(out)

    out.seek(0)

    return out


